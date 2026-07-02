from __future__ import annotations

import json
import sys
import tkinter as tk
from pathlib import Path
from tkinter import messagebox, ttk

ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

BG = "#f5f7fb"
PANEL = "#ffffff"
PANEL_2 = "#f8fafc"
CARD = "#ffffff"
ACCENT = "#2563eb"
TEXT = "#0f172a"
MUTED = "#64748b"
BORDER = "#dbe3f0"
ERROR = "#e11d48"
LIST_SELECT = "#e8f0ff"
LIST_SELECT_TEXT = "#0f172a"
FONT_UI = "Segoe UI"
FONT_MONO = "Consolas"


_FETCH_MOD = None
_FUNDI_MOD = None
_OPENPYXL_LOAD_WORKBOOK = None


def fetch_mod():
    global _FETCH_MOD
    if _FETCH_MOD is None:
        import fetch_med_lit as mod

        _FETCH_MOD = mod
    return _FETCH_MOD


def fundi_mod():
    global _FUNDI_MOD
    if _FUNDI_MOD is None:
        import fundigraph_review_workflow as mod

        _FUNDI_MOD = mod
    return _FUNDI_MOD


def load_workbook_fn():
    global _OPENPYXL_LOAD_WORKBOOK
    if _OPENPYXL_LOAD_WORKBOOK is None:
        from openpyxl import load_workbook as fn

        _OPENPYXL_LOAD_WORKBOOK = fn
    return _OPENPYXL_LOAD_WORKBOOK


class ScrollableDetailMixin:
    def init_scroll_state(self) -> None:
        self.right_shell: tk.Frame | None = None
        self.detail_canvas: tk.Canvas | None = None
        self._detail_window = None
        self._wheel_remainder = 0.0

    def bind_scroll_events(self) -> None:
        self.winfo_toplevel().bind_all("<MouseWheel>", self._on_mousewheel, add="+")
        self.winfo_toplevel().bind_all("<Shift-MouseWheel>", self._on_mousewheel_horizontal, add="+")
        self.winfo_toplevel().bind_all("<Button-4>", self._on_button4_scroll, add="+")
        self.winfo_toplevel().bind_all("<Button-5>", self._on_button5_scroll, add="+")

    def _is_inside_right_panel(self, widget: tk.Widget | None) -> bool:
        if widget is None or self.right_shell is None:
            return False
        current = widget
        while current is not None:
            if current == self.right_shell:
                return True
            parent_name = current.winfo_parent()
            if not parent_name:
                break
            try:
                current = current.nametowidget(parent_name)
            except Exception:
                break
        return False

    def _on_detail_frame_configure(self, _event=None) -> None:
        if self.detail_canvas is not None:
            self.detail_canvas.configure(scrollregion=self.detail_canvas.bbox("all"))

    def _on_detail_canvas_configure(self, event) -> None:
        if self.detail_canvas is not None and self._detail_window is not None:
            self.detail_canvas.itemconfigure(self._detail_window, width=event.width)

    def _scroll_detail_units(self, units: int) -> str | None:
        if self.detail_canvas is None or units == 0:
            return None
        self.detail_canvas.yview_scroll(units, "units")
        return "break"

    def _on_mousewheel(self, event) -> str | None:
        widget = self.winfo_toplevel().winfo_containing(event.x_root, event.y_root)
        if not self._is_inside_right_panel(widget):
            return None
        delta = getattr(event, "delta", 0)
        if not delta:
            return None
        steps = delta / 120.0
        if abs(steps) < 1:
            steps = 0.35 if steps > 0 else -0.35
        self._wheel_remainder += steps
        whole_units = int(self._wheel_remainder)
        if whole_units == 0:
            whole_units = 1 if self._wheel_remainder > 0 else -1
            self._wheel_remainder = 0.0
        else:
            self._wheel_remainder -= whole_units
        return self._scroll_detail_units(-whole_units)

    def _on_mousewheel_horizontal(self, event) -> str | None:
        return self._on_mousewheel(event)

    def _on_button4_scroll(self, event) -> str | None:
        widget = self.winfo_toplevel().winfo_containing(event.x_root, event.y_root)
        if not self._is_inside_right_panel(widget):
            return None
        return self._scroll_detail_units(-1)

    def _on_button5_scroll(self, event) -> str | None:
        widget = self.winfo_toplevel().winfo_containing(event.x_root, event.y_root)
        if not self._is_inside_right_panel(widget):
            return None
        return self._scroll_detail_units(1)


class LiteratureFrame(tk.Frame, ScrollableDetailMixin):
    def __init__(self, master: tk.Widget) -> None:
        tk.Frame.__init__(self, master, bg=BG)
        self.init_scroll_state()

        self.diseases: list[str] = []
        self.store: dict = {}
        self.review_payload: dict = {}
        self.results_rows: list[dict[str, str]] = []
        self.no_result_rows: dict[str, dict[str, str]] = {}
        self.disease_items: dict[str, dict] = {}
        self.visible_keys: list[str] = []
        self.current_key: str | None = None

        self.summary_var = tk.StringVar()
        self.review_meta_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Select a disease to review.")
        self.filter_var = tk.StringVar(value="all")
        self.search_var = tk.StringVar()
        self.strategy_status_var = tk.StringVar(value="pending")

        self.disease_list: tk.Listbox | None = None
        self.overview_text: tk.Text | None = None
        self.term_text: tk.Text | None = None
        self.query_text: tk.Text | None = None
        self.notes_text: tk.Text | None = None
        self.comment_text: tk.Text | None = None
        self.results_text: tk.Text | None = None
        self.no_results_text: tk.Text | None = None

        self._build_ui()
        self.bind_scroll_events()
        self.reload_all()

    def _build_ui(self) -> None:
        top = ttk.Frame(self, style="App.TFrame", padding=18)
        top.pack(fill="x")
        ttk.Label(top, text="Literature Search Desk", style="Title.TLabel").pack(anchor="w")
        ttk.Label(
            top,
            text="Review diseases, stored search strategies, and literature results without editing raw JSON.",
            style="Sub.TLabel",
        ).pack(anchor="w", pady=(4, 14))

        toolbar = ttk.Frame(top, style="App.TFrame")
        toolbar.pack(fill="x")
        ttk.Button(toolbar, text="Reload", style="Secondary.TButton", command=self.reload_all).pack(side="left")
        ttk.Button(toolbar, text="Generate Missing", style="Accent.TButton", command=self.generate_missing).pack(
            side="left", padx=(10, 0)
        )
        ttk.Button(toolbar, text="Save Current", style="Secondary.TButton", command=self.save_current_item).pack(
            side="left", padx=(10, 0)
        )
        ttk.Button(toolbar, text="Save All", style="Primary.TButton", command=self.save_all).pack(side="left", padx=(10, 0))
        ttk.Label(toolbar, text="Filter", style="Sub.TLabel").pack(side="left", padx=(18, 8))
        filter_box = ttk.Combobox(
            toolbar,
            textvariable=self.filter_var,
            values=["all", "approved", "pending", "rejected", "missing"],
            state="readonly",
            width=12,
            style="App.TCombobox",
        )
        filter_box.pack(side="left")
        filter_box.bind("<<ComboboxSelected>>", lambda _e: self.refresh_disease_list())

        search_entry = tk.Entry(toolbar, textvariable=self.search_var, bd=0, relief="flat", font=(FONT_UI, 10))
        search_entry.pack(side="left", padx=(12, 0), ipadx=90, ipady=6)
        search_entry.bind("<KeyRelease>", lambda _e: self.refresh_disease_list())
        ttk.Label(toolbar, textvariable=self.summary_var, style="Sub.TLabel").pack(side="right")

        body = ttk.Frame(self, style="App.TFrame", padding=(18, 0, 18, 18))
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=0)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        left = tk.Frame(body, bg=PANEL, highlightbackground=BORDER, highlightthickness=1)
        left.grid(row=0, column=0, sticky="nsw", padx=(0, 14))
        left.configure(width=330)
        left.grid_propagate(False)

        self.right_shell = tk.Frame(body, bg=PANEL, highlightbackground=BORDER, highlightthickness=1)
        self.right_shell.grid(row=0, column=1, sticky="nsew")
        self.right_shell.grid_columnconfigure(0, weight=1)
        self.right_shell.grid_rowconfigure(0, weight=1)

        self.detail_canvas = tk.Canvas(self.right_shell, bg=PANEL, highlightthickness=0, bd=0, relief="flat")
        detail_scrollbar = ttk.Scrollbar(self.right_shell, orient="vertical", command=self.detail_canvas.yview)
        self.detail_canvas.configure(yscrollcommand=detail_scrollbar.set)
        self.detail_canvas.grid(row=0, column=0, sticky="nsew")
        detail_scrollbar.grid(row=0, column=1, sticky="ns")

        right = tk.Frame(self.detail_canvas, bg=PANEL)
        right.grid_columnconfigure(0, weight=1)
        self._detail_window = self.detail_canvas.create_window((0, 0), window=right, anchor="nw")
        right.bind("<Configure>", self._on_detail_frame_configure)
        self.detail_canvas.bind("<Configure>", self._on_detail_canvas_configure)

        ttk.Label(left, text="Diseases", style="Section.TLabel").pack(anchor="w", padx=14, pady=(14, 8))
        self.disease_list = tk.Listbox(
            left,
            bg=PANEL,
            fg=TEXT,
            selectbackground=LIST_SELECT,
            selectforeground=LIST_SELECT_TEXT,
            activestyle="none",
            relief="flat",
            highlightthickness=0,
            bd=0,
            font=(FONT_UI, 10),
        )
        self.disease_list.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.disease_list.bind("<<ListboxSelect>>", self.on_select_disease)

        header = tk.Frame(right, bg=PANEL)
        header.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 10))
        header.grid_columnconfigure(0, weight=1)
        tk.Label(header, textvariable=self.status_var, bg=PANEL, fg=TEXT, font=(FONT_UI, 14, "bold")).grid(
            row=0, column=0, sticky="w"
        )

        status_row = tk.Frame(right, bg=PANEL)
        status_row.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 10))
        tk.Label(status_row, text="Strategy Status", bg=PANEL, fg=TEXT, font=(FONT_UI, 10, "bold")).pack(side="left")
        status_box = ttk.Combobox(
            status_row,
            textvariable=self.strategy_status_var,
            values=["pending", "approved", "rejected"],
            state="readonly",
            width=12,
            style="App.TCombobox",
        )
        status_box.pack(side="left", padx=(12, 0))

        self.overview_text = self._text_block(right, "Overview", row=2, height=7, readonly=True)
        self.term_text = self._text_block(right, "OpenAlex Terms (one per line)", row=3, height=7)
        self.query_text = self._text_block(right, "Europe PMC Query", row=4, height=5)
        self.notes_text = self._text_block(right, "Notes (one per line)", row=5, height=6)
        self.comment_text = self._text_block(right, "Review Comment", row=6, height=4)

        actions = tk.Frame(right, bg=PANEL)
        actions.grid(row=7, column=0, sticky="ew", padx=16, pady=(4, 10))
        ttk.Button(actions, text="Save Current", style="Secondary.TButton", command=self.save_current_item).pack(side="left")
        ttk.Button(actions, text="Save All", style="Primary.TButton", command=self.save_all).pack(side="left", padx=(10, 0))

        self.results_text = self._text_block(right, "Literature Results", row=8, height=12, readonly=True)
        self.no_results_text = self._text_block(right, "No-Result Summary", row=9, height=6, readonly=True)

    def _text_block(self, parent: tk.Widget, title: str, row: int, height: int, readonly: bool = False) -> tk.Text:
        wrap = tk.Frame(parent, bg=PANEL)
        wrap.grid(row=row, column=0, sticky="ew", padx=16, pady=(0, 10))
        tk.Label(wrap, text=title, bg=PANEL, fg=TEXT, font=(FONT_UI, 10, "bold")).pack(anchor="w", pady=(0, 6))
        text = tk.Text(
            wrap,
            height=height,
            wrap="word",
            bg=CARD,
            fg=TEXT,
            insertbackground=TEXT,
            selectbackground="#dbeafe",
            relief="flat",
            bd=0,
            padx=12,
            pady=12,
            highlightbackground=BORDER,
            highlightthickness=1,
            font=(FONT_MONO, 10),
        )
        text.pack(fill="both", expand=True)
        if readonly:
            text.configure(state="disabled")
        return text

    def reload_all(self) -> None:
        mod = fetch_mod()
        self.diseases = mod.read_diseases(mod.DISEASE_FILE) if mod.DISEASE_FILE.exists() else []
        self.store = mod.load_store()
        self.review_payload = mod.load_json_file(
            mod.SEARCH_REVIEW_FILE,
            {"schema_version": 1, "generated_on": mod.today_iso(), "review_items": []},
        )
        self.results_rows, self.no_result_rows = self._load_workbook_data()
        self._build_disease_items()
        self.refresh_disease_list()

    def _load_workbook_data(self) -> tuple[list[dict[str, str]], dict[str, dict[str, str]]]:
        mod = fetch_mod()
        if not mod.DOI_FILE.exists():
            return [], {}
        wb = load_workbook_fn()(mod.DOI_FILE, read_only=True, data_only=True)
        results_rows: list[dict[str, str]] = []
        no_result_rows: dict[str, dict[str, str]] = {}
        if "Results" in wb.sheetnames:
            ws = wb["Results"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(cell) if cell is not None else "" for cell in rows[0]]
                for row in rows[1:]:
                    if not row or not any(cell is not None for cell in row):
                        continue
                    results_rows.append({header[i]: ("" if row[i] is None else str(row[i])) for i in range(min(len(header), len(row)))})
        if "No_Result_Summary" in wb.sheetnames:
            ws = wb["No_Result_Summary"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(cell) if cell is not None else "" for cell in rows[0]]
                for row in rows[1:]:
                    if not row or not row[0]:
                        continue
                    item = {header[i]: ("" if row[i] is None else str(row[i])) for i in range(min(len(header), len(row)))}
                    no_result_rows[item["Disease"]] = item
        return results_rows, no_result_rows

    def _build_disease_items(self) -> None:
        self.disease_items = {}
        approved = self.store.get("strategies", {})
        pending_list = self.review_payload.get("review_items", [])
        pending_map = {item["canonical_key"]: item for item in pending_list}
        names = set(self.diseases)
        for item in approved.values():
            names.add(item.get("disease_name", ""))
        for item in pending_list:
            names.add(item.get("disease_name", ""))
        for row in self.results_rows:
            names.add(row.get("Disease", ""))
        for name in self.no_result_rows:
            names.add(name)
        for disease in sorted(name for name in names if name):
            generated = fetch_mod().generate_search_terms(disease)
            key = generated["canonical_key"]
            strategy = approved.get(key)
            review_item = pending_map.get(key)
            if strategy:
                status = strategy.get("status", "approved")
            elif review_item:
                status = review_item.get("status", "pending")
            else:
                status = "missing"
            self.disease_items[key] = {
                "disease_name": disease,
                "canonical_key": key,
                "status": status,
                "strategy": strategy,
                "review_item": review_item,
                "generated": generated,
            }

    def refresh_disease_list(self) -> None:
        if self.disease_list is None:
            return
        self.disease_list.delete(0, tk.END)
        self.visible_keys = []
        query = self.search_var.get().strip().lower()
        mode = self.filter_var.get()
        for key, item in self.disease_items.items():
            disease = item["disease_name"]
            status = item["status"]
            if query and query not in disease.lower():
                continue
            if mode != "all" and status != mode:
                continue
            self.visible_keys.append(key)
            self.disease_list.insert(tk.END, f"{self._status_badge(status)}  {disease}")
        approved = sum(1 for item in self.disease_items.values() if item["status"] == "approved")
        pending = sum(1 for item in self.disease_items.values() if item["status"] == "pending")
        missing = sum(1 for item in self.disease_items.values() if item["status"] == "missing")
        self.summary_var.set(f"Approved {approved}   Pending {pending}   Missing {missing}")
        if self.visible_keys:
            self.disease_list.selection_clear(0, tk.END)
            self.disease_list.selection_set(0)
            self.show_disease(self.visible_keys[0])
        else:
            self.status_var.set("No disease matches the current filter.")
            self._fill_text(self.overview_text, "", readonly=True)
            self._fill_text(self.term_text, "")
            self._fill_text(self.query_text, "")
            self._fill_text(self.notes_text, "")
            self._fill_text(self.comment_text, "")
            self._fill_text(self.results_text, "", readonly=True)
            self._fill_text(self.no_results_text, "", readonly=True)

    def _status_badge(self, status: str) -> str:
        return {"approved": "APP", "pending": "PND", "rejected": "REJ", "missing": "NEW"}.get(status, "UNK")

    def on_select_disease(self, _event=None) -> None:
        if self.disease_list is None:
            return
        selection = self.disease_list.curselection()
        if not selection:
            return
        self.show_disease(self.visible_keys[selection[0]])

    def show_disease(self, key: str) -> None:
        self.current_key = key
        item = self.disease_items[key]
        disease = item["disease_name"]
        self.status_var.set(f"{disease}   {item['status'].upper()}")
        strategy = item["strategy"] or item["review_item"]
        generated = item["generated"]
        terms = (strategy or {}).get("queries", {}).get("openalex_terms", generated["queries"]["openalex_terms"])
        query = (strategy or {}).get("queries", {}).get("europe_pmc_query", generated["queries"]["europe_pmc_query"])
        notes = (strategy or {}).get("notes", generated["notes"])
        comment = (strategy or {}).get("review_comment", "")
        status = (strategy or {}).get("status", item["status"])
        self.strategy_status_var.set(status if status in {"pending", "approved", "rejected"} else "pending")
        overview_lines = [
            f"Disease: {disease}",
            f"Canonical key: {generated['canonical_key']}",
            f"Current status: {status}",
            f"Stored strategy: {'Yes' if item['strategy'] else 'No'}",
            f"Pending review item: {'Yes' if item['review_item'] else 'No'}",
            f"Literature rows: {sum(1 for row in self.results_rows if row.get('Disease') == disease)}",
            f"No-result summary: {'Yes' if disease in self.no_result_rows else 'No'}",
        ]
        self._fill_text(self.overview_text, "\n".join(overview_lines), readonly=True)
        self._fill_text(self.term_text, "\n".join(terms))
        self._fill_text(self.query_text, query)
        self._fill_text(self.notes_text, "\n".join(notes))
        self._fill_text(self.comment_text, comment)
        result_lines = []
        for row in self.results_rows:
            if row.get("Disease") != disease:
                continue
            result_lines.append(
                "\n".join(
                    [
                        f"Title: {row.get('Title', '')}",
                        f"Date: {row.get('Publication Date', '')}",
                        f"Type: {row.get('Type', '')}",
                        f"Source: {row.get('Source Database', '')}",
                        f"Identifier: {row.get('Identifier', '') or row.get('DOI', '')}",
                        f"Landing Page: {row.get('Landing Page', '')}",
                        f"Abstract File: {row.get('Abstract File', '')}",
                    ]
                )
            )
        self._fill_text(self.results_text, "\n\n---\n\n".join(result_lines) if result_lines else "No stored result rows.", readonly=True)
        no_row = self.no_result_rows.get(disease)
        no_text = ""
        if no_row:
            no_text = "\n".join(
                [
                    f"Status: {no_row.get('Status', '')}",
                    f"Date Filter: {no_row.get('Date Filter', '')}",
                    f"Checked Sources: {no_row.get('Checked Sources', '')}",
                ]
            )
        self._fill_text(self.no_results_text, no_text or "No no-result summary for this disease.", readonly=True)

    def _fill_text(self, widget: tk.Text | None, value: str, readonly: bool = False) -> None:
        if widget is None:
            return
        widget.configure(state="normal")
        widget.delete("1.0", tk.END)
        widget.insert("1.0", value)
        if readonly:
            widget.configure(state="disabled")

    def _build_payload_from_editor(self) -> dict | None:
        if not self.current_key or self.term_text is None or self.query_text is None or self.notes_text is None or self.comment_text is None:
            return None
        item = self.disease_items[self.current_key]
        payload = {
            "disease_name": item["disease_name"],
            "canonical_key": item["canonical_key"],
            "status": self.strategy_status_var.get().strip() or "pending",
            "generated_on": fetch_mod().today_iso(),
            "notes": [line.strip() for line in self.notes_text.get("1.0", tk.END).splitlines() if line.strip()],
            "queries": {
                "openalex_terms": [line.strip() for line in self.term_text.get("1.0", tk.END).splitlines() if line.strip()],
                "europe_pmc_query": self.query_text.get("1.0", tk.END).strip(),
            },
            "review_comment": self.comment_text.get("1.0", tk.END).strip(),
        }
        if payload["status"] == "approved":
            payload["approved_on"] = fetch_mod().today_iso()
        return payload

    def _merge_payload(self, payload: dict) -> None:
        strategies = self.store.setdefault("strategies", {})
        review_items = self.review_payload.setdefault("review_items", [])
        key = payload["canonical_key"]
        existing_index = next((i for i, item in enumerate(review_items) if item.get("canonical_key") == key), None)
        if existing_index is None:
            review_items.append(payload)
        else:
            review_items[existing_index] = payload
        if payload["status"] == "approved":
            strategies[key] = payload
        elif key in strategies:
            del strategies[key]
        self.store["generated_on"] = fetch_mod().today_iso()
        self.review_payload["generated_on"] = fetch_mod().today_iso()

    def save_current_item(self) -> None:
        payload = self._build_payload_from_editor()
        if not payload:
            messagebox.showwarning("No Selection", "Select a disease first.")
            return
        self._merge_payload(payload)
        self._save_review_files()
        self.reload_all()
        messagebox.showinfo("Saved", f"Saved strategy for {payload['disease_name']}.")

    def save_all(self) -> None:
        if self.current_key:
            payload = self._build_payload_from_editor()
            if payload:
                self._merge_payload(payload)
        self._save_review_files()
        self.reload_all()
        messagebox.showinfo("Saved", "Saved strategy files.")

    def _save_review_files(self) -> None:
        self.review_payload.setdefault(
            "instructions",
            [
                "Review each pending disease strategy.",
                "Change status from pending to approved or rejected.",
                "You may edit the query terms before approval.",
                "After approval, rerun the retrieval workflow and the approved strategy will move into the local store automatically.",
            ],
        )
        mod = fetch_mod()
        mod.save_json_file(mod.SEARCH_STORE_FILE, self.store)
        mod.save_json_file(mod.SEARCH_REVIEW_FILE, self.review_payload)
        lines = ["# Search Strategy Review", "", "Review the candidate search strategies below.", ""]
        for index, item in enumerate(self.review_payload.get("review_items", []), start=1):
            lines.extend(
                [
                    f"## {index}. {item.get('disease_name', '')}",
                    "",
                    f"- Canonical key: `{item.get('canonical_key', '')}`",
                    f"- Status: `{item.get('status', '')}`",
                    "- OpenAlex terms:",
                    f"  - `{', '.join(item.get('queries', {}).get('openalex_terms', []))}`",
                    "- Europe PMC query:",
                    f"  - `{item.get('queries', {}).get('europe_pmc_query', '')}`",
                    "",
                ]
            )
        fetch_mod().SEARCH_REVIEW_MARKDOWN.write_text("\n".join(lines), encoding="utf-8")

    def generate_missing(self) -> None:
        mod = fetch_mod()
        pending_items = mod.ensure_strategy_candidates(self.diseases, self.store)
        if not pending_items:
            messagebox.showinfo("No New Diseases", "No new diseases need a fresh candidate strategy.")
            return
        existing = {item.get("canonical_key"): item for item in self.review_payload.get("review_items", [])}
        for item in pending_items:
            existing[item["canonical_key"]] = item
        self.review_payload["review_items"] = list(existing.values())
        self.review_payload["generated_on"] = mod.today_iso()
        self._save_review_files()
        self.reload_all()
        messagebox.showinfo("Generated", f"Generated {len(pending_items)} candidate strategies.")


class FundiGraphFrame(tk.Frame, ScrollableDetailMixin):
    def __init__(self, master: tk.Widget) -> None:
        tk.Frame.__init__(self, master, bg=BG)
        self.init_scroll_state()

        mod = fundi_mod()
        self.payload = mod.build_review_payload()
        self.reviewed_path = mod.reviewed_json_path()
        self.state = self._build_state()
        self.current_index = 0
        self.filter_mode = tk.StringVar(value="pending")
        self.visible_indices: list[int] = []

        self.listbox: tk.Listbox | None = None
        self.meta: tk.Label | None = None
        self.evidence: tk.Text | None = None
        self.original_triple: tk.Text | None = None
        self.corrected: tk.Text | None = None
        self.status_var = tk.StringVar()
        self.position_var = tk.StringVar()
        self.summary_var = tk.StringVar()
        self.review_meta_var = tk.StringVar()

        self._build_ui()
        self.bind_scroll_events()
        self._load_existing_review()
        self._refresh_list()
        if self.visible_indices:
            self._show_item(self.visible_indices[0])
        elif self.state:
            self._show_item(0)

    def _build_state(self) -> list[dict]:
        return [
            {
                "id": item["id"],
                "decision": "pending",
                "corrected_triple": item["triple"],
                "original": item,
            }
            for item in self.payload["items"]
        ]

    def _build_ui(self) -> None:
        top = ttk.Frame(self, style="App.TFrame", padding=18)
        top.pack(fill="x")
        ttk.Label(top, text="FundiGraph Review", style="Title.TLabel").pack(anchor="w")
        ttk.Label(
            top,
            text="Review extracted triplets and apply verified changes into FundiGraph.",
            style="Sub.TLabel",
        ).pack(anchor="w", pady=(4, 14))
        tk.Label(
            top,
            textvariable=self.review_meta_var,
            bg=BG,
            fg=MUTED,
            font=(FONT_UI, 10),
            anchor="w",
            justify="left",
        ).pack(anchor="w", pady=(0, 12))

        toolbar = ttk.Frame(top, style="App.TFrame")
        toolbar.pack(fill="x")
        ttk.Button(toolbar, text="Save Review", style="Secondary.TButton", command=self.save_review).pack(side="left")
        ttk.Button(toolbar, text="Save And Apply", style="Primary.TButton", command=self.save_and_apply).pack(
            side="left", padx=(10, 0)
        )
        ttk.Button(toolbar, text="Copy HTML Path", style="Accent.TButton", command=self.copy_html_path).pack(
            side="left", padx=(10, 0)
        )
        ttk.Label(toolbar, text="Filter", style="Sub.TLabel").pack(side="left", padx=(20, 8))
        filter_box = ttk.Combobox(
            toolbar,
            textvariable=self.filter_mode,
            values=["pending", "new_only", "error", "confirmed", "all"],
            state="readonly",
            width=12,
            style="App.TCombobox",
        )
        filter_box.pack(side="left")
        filter_box.bind("<<ComboboxSelected>>", lambda _e: self.on_filter_change())
        ttk.Label(toolbar, textvariable=self.summary_var, style="Sub.TLabel").pack(side="right")

        body = ttk.Frame(self, style="App.TFrame", padding=(18, 0, 18, 18))
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=0)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        left = tk.Frame(body, bg=PANEL, highlightbackground=BORDER, highlightthickness=1)
        left.grid(row=0, column=0, sticky="nsw", padx=(0, 14))
        left.configure(width=340)
        left.grid_propagate(False)

        self.right_shell = tk.Frame(body, bg=PANEL, highlightbackground=BORDER, highlightthickness=1)
        self.right_shell.grid(row=0, column=1, sticky="nsew")
        self.right_shell.grid_columnconfigure(0, weight=1)
        self.right_shell.grid_rowconfigure(0, weight=1)

        self.detail_canvas = tk.Canvas(self.right_shell, bg=PANEL, highlightthickness=0, bd=0, relief="flat")
        detail_scrollbar = ttk.Scrollbar(self.right_shell, orient="vertical", command=self.detail_canvas.yview)
        self.detail_canvas.configure(yscrollcommand=detail_scrollbar.set)
        self.detail_canvas.grid(row=0, column=0, sticky="nsew")
        detail_scrollbar.grid(row=0, column=1, sticky="ns")

        right = tk.Frame(self.detail_canvas, bg=PANEL)
        right.grid_columnconfigure(0, weight=1)
        self._detail_window = self.detail_canvas.create_window((0, 0), window=right, anchor="nw")
        right.bind("<Configure>", self._on_detail_frame_configure)
        self.detail_canvas.bind("<Configure>", self._on_detail_canvas_configure)

        ttk.Label(left, text="Queue", style="Section.TLabel").pack(anchor="w", padx=14, pady=(14, 8))
        self.listbox = tk.Listbox(
            left,
            bg=PANEL,
            fg=TEXT,
            selectbackground=LIST_SELECT,
            selectforeground=LIST_SELECT_TEXT,
            activestyle="none",
            relief="flat",
            highlightthickness=0,
            bd=0,
            font=(FONT_UI, 10),
        )
        self.listbox.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.listbox.bind("<<ListboxSelect>>", self.on_select)

        header = tk.Frame(right, bg=PANEL)
        header.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 10))
        header.grid_columnconfigure(0, weight=1)
        tk.Label(header, textvariable=self.status_var, bg=PANEL, fg=TEXT, font=(FONT_UI, 14, "bold")).grid(
            row=0, column=0, sticky="w"
        )
        tk.Label(header, textvariable=self.position_var, bg=PANEL, fg=ACCENT, font=(FONT_UI, 10)).grid(
            row=0, column=1, sticky="e"
        )

        nav = tk.Frame(right, bg=PANEL)
        nav.grid(row=1, column=0, sticky="ew", padx=16)
        ttk.Button(nav, text="Previous", style="Secondary.TButton", command=self.show_previous).pack(side="left")
        ttk.Button(nav, text="Next", style="Secondary.TButton", command=self.show_next).pack(side="left", padx=(10, 0))
        tk.Label(nav, text="Confirm/Error will auto-advance", bg=PANEL, fg=MUTED, font=(FONT_UI, 9)).pack(side="right")

        self.meta = tk.Label(
            right,
            bg=CARD,
            fg=MUTED,
            justify="left",
            anchor="nw",
            padx=14,
            pady=12,
            font=(FONT_UI, 10),
            highlightbackground=BORDER,
            highlightthickness=1,
        )
        self.meta.grid(row=2, column=0, sticky="ew", padx=16, pady=(12, 10))

        self.evidence = self._text_block(right, "Evidence", row=3, height=5, readonly=True)
        self.original_triple = self._text_block(right, "Original Triple", row=4, height=5, readonly=True)

        actions = tk.Frame(right, bg=PANEL)
        actions.grid(row=5, column=0, sticky="ew", padx=16, pady=(4, 10))
        ttk.Button(actions, text="Confirm", style="Accent.TButton", command=lambda: self.set_decision("confirm")).pack(
            side="left"
        )
        ttk.Button(actions, text="Error", style="Danger.TButton", command=lambda: self.set_decision("error")).pack(
            side="left", padx=(10, 0)
        )
        ttk.Button(actions, text="Reset", style="Secondary.TButton", command=lambda: self.set_decision("pending", False)).pack(
            side="left", padx=(10, 0)
        )

        self.corrected = self._text_block(
            right,
            "Corrected Triple",
            row=6,
            height=14,
            readonly=False,
            helper_text="Edit only when the extracted relation needs manual correction.",
        )

    def _text_block(
        self,
        parent: tk.Widget,
        title: str,
        row: int,
        height: int,
        readonly: bool,
        helper_text: str = "",
    ) -> tk.Text:
        wrap = tk.Frame(parent, bg=PANEL)
        wrap.grid(row=row, column=0, sticky="ew", padx=16, pady=(0, 10))
        tk.Label(wrap, text=title, bg=PANEL, fg=TEXT, font=(FONT_UI, 10, "bold")).pack(anchor="w", pady=(0, 6))
        if helper_text:
            tk.Label(wrap, text=helper_text, bg=PANEL, fg=MUTED, font=(FONT_UI, 9)).pack(anchor="w", pady=(0, 8))
        text = tk.Text(
            wrap,
            height=height,
            wrap="word",
            bg=CARD,
            fg=TEXT,
            insertbackground=TEXT,
            selectbackground="#dbeafe",
            relief="flat",
            bd=0,
            padx=14,
            pady=14,
            highlightbackground=BORDER,
            highlightthickness=1,
            font=(FONT_MONO, 10),
        )
        text.pack(fill="both", expand=True)
        if readonly:
            text.configure(state="disabled")
        return text

    def _load_existing_review(self) -> None:
        if not self.reviewed_path.exists():
            return
        try:
            data = json.loads(self.reviewed_path.read_text(encoding="utf-8"))
        except Exception:
            return
        by_id = {item["id"]: item for item in data.get("items", [])}
        for row in self.state:
            saved = by_id.get(row["id"])
            if not saved:
                continue
            row["decision"] = saved.get("decision", "pending")
            row["corrected_triple"] = saved.get("corrected_triple", row["corrected_triple"])

    def _row_tint(self, row: dict) -> str:
        if row["decision"] == "confirm":
            return "ok"
        if row["decision"] == "error":
            return "fix"
        if row["original"]["planned_action"] != "skip":
            return "new"
        return "hold"

    def _refresh_list(self) -> None:
        if self.listbox is None:
            return
        self._persist_current_text()
        self.listbox.delete(0, tk.END)
        self.visible_indices = []
        for index, row in enumerate(self.state):
            if not self._matches_filter(row):
                continue
            self.visible_indices.append(index)
            self.listbox.insert(tk.END, f"{row['id']:02d}  {self._row_tint(row).upper():<4}  {row['original']['relation']}")
        confirm = sum(1 for row in self.state if row["decision"] == "confirm")
        error = sum(1 for row in self.state if row["decision"] == "error")
        pending = sum(1 for row in self.state if row["decision"] == "pending")
        self.summary_var.set(
            f"Confirmed {confirm}   Error {error}   Pending {pending}   Visible {len(self.visible_indices)}/{len(self.state)}"
        )
        self.review_meta_var.set(
            "Extracted {extracted}   Dropped overlap>=0.80 {dropped}   Gemini queue {queued}   "
            "Gemini Correct/Error {correct}/{wrong}".format(
                extracted=self.payload.get("extracted_candidate_count", len(self.state)),
                dropped=self.payload.get("overlap_dropped_count", 0),
                queued=self.payload.get("review_candidate_count", len(self.state)),
                correct=self.payload.get("gemini_precheck", {}).get("summary", {}).get("correct_count", 0),
                wrong=self.payload.get("gemini_precheck", {}).get("summary", {}).get("error_count", 0),
            )
        )
        if not self.visible_indices:
            self._clear_detail("No items match the current filter.")
            return
        if self.current_index not in self.visible_indices:
            self._show_item(self.visible_indices[0])
        else:
            self._sync_list_selection()
            self._show_item(self.current_index, update_selection=False)

    def _matches_filter(self, row: dict) -> bool:
        mode = self.filter_mode.get()
        if mode == "pending":
            return row["decision"] == "pending"
        if mode == "confirmed":
            return row["decision"] == "confirm"
        if mode == "error":
            return row["decision"] == "error"
        if mode == "new_only":
            return row["original"]["planned_action"] != "skip"
        return True

    def _clear_detail(self, message: str) -> None:
        self.status_var.set(message)
        self.position_var.set("0 / 0")
        if self.meta is not None:
            self.meta.configure(text="")
        self._set_text(self.evidence, "")
        self._set_text(self.original_triple, "")
        if self.corrected is not None:
            self.corrected.configure(state="normal")
            self.corrected.delete("1.0", tk.END)

    def _sync_list_selection(self) -> None:
        if self.listbox is None or self.current_index not in self.visible_indices:
            return
        visible_pos = self.visible_indices.index(self.current_index)
        self.listbox.selection_clear(0, tk.END)
        self.listbox.selection_set(visible_pos)
        self.listbox.activate(visible_pos)

    def _set_text(self, widget: tk.Text | None, value: str) -> None:
        if widget is None:
            return
        widget.configure(state="normal")
        widget.delete("1.0", tk.END)
        widget.insert("1.0", value)
        if widget is not self.corrected:
            widget.configure(state="disabled")

    def _show_item(self, index: int, update_selection: bool = True) -> None:
        if not self.state:
            return
        index = max(0, min(index, len(self.state) - 1))
        self.current_index = index
        row = self.state[index]
        item = row["original"]
        self.status_var.set(f"Item {row['id']}   {row['decision'].upper()}   {item['relation']}")
        if self.meta is not None:
            self.meta.configure(
                text="\n".join(
                        [
                            f"Source: {item['source_file']}",
                            f"Confidence: {item['confidence']:.2f}",
                            f"Overlap: {item['overlap']['matched']}",
                            f"Best Match: {item['overlap']['value'] or 'N/A'}",
                            f"Gemini审核结果: {item.get('gemini_precheck', {}).get('verdict', 'unavailable').capitalize()}",
                            f"Planned Action: {item['planned_action']}",
                            f"Gemini Reason: {item.get('gemini_precheck', {}).get('reason', '')}",
                        ]
                    )
            )
        self._set_text(self.evidence, item["source_excerpt"])
        self._set_text(self.original_triple, item["triple"])
        if self.corrected is not None:
            self.corrected.configure(state="normal")
            self.corrected.delete("1.0", tk.END)
            self.corrected.insert("1.0", row["corrected_triple"])
        if self.visible_indices and index in self.visible_indices:
            self.position_var.set(f"{self.visible_indices.index(index) + 1} / {len(self.visible_indices)}")
        else:
            self.position_var.set("0 / 0")
        if update_selection:
            self._sync_list_selection()

    def on_select(self, _event=None) -> None:
        if self.listbox is None:
            return
        selection = self.listbox.curselection()
        if not selection or not self.visible_indices:
            return
        self._persist_current_text()
        self._show_item(self.visible_indices[selection[0]])

    def _persist_current_text(self) -> None:
        if not self.state or self.corrected is None:
            return
        self.state[self.current_index]["corrected_triple"] = self.corrected.get("1.0", tk.END).strip()

    def set_decision(self, decision: str, advance: bool = True) -> None:
        self._persist_current_text()
        self.state[self.current_index]["decision"] = decision
        previous_index = self.current_index
        next_index = self._next_visible_index(previous_index) if advance and decision in {"confirm", "error"} else previous_index
        self._refresh_list()
        if self.visible_indices:
            target = next_index if next_index in self.visible_indices else self.visible_indices[0]
            self._show_item(target)

    def _next_visible_index(self, current: int) -> int:
        if current not in self.visible_indices:
            return self.visible_indices[0] if self.visible_indices else current
        pos = self.visible_indices.index(current)
        if pos + 1 < len(self.visible_indices):
            return self.visible_indices[pos + 1]
        return current

    def on_filter_change(self) -> None:
        self._refresh_list()

    def show_previous(self) -> None:
        if not self.visible_indices:
            return
        if self.current_index not in self.visible_indices:
            self._show_item(self.visible_indices[0])
            return
        pos = self.visible_indices.index(self.current_index)
        self._persist_current_text()
        self._show_item(self.visible_indices[max(0, pos - 1)])

    def show_next(self) -> None:
        if not self.visible_indices:
            return
        if self.current_index not in self.visible_indices:
            self._show_item(self.visible_indices[0])
            return
        pos = self.visible_indices.index(self.current_index)
        self._persist_current_text()
        self._show_item(self.visible_indices[min(len(self.visible_indices) - 1, pos + 1)])

    def save_review(self) -> None:
        self._persist_current_text()
        data = {
            "generated_on": self.payload["generated_on"],
            "source_review_json": str(fundi_mod().review_html_path().name),
            "items": [
                {
                    "id": row["id"],
                    "decision": row["decision"],
                    "corrected_triple": row["corrected_triple"],
                    "source_file": row["original"]["source_file"],
                    "source_excerpt": row["original"]["source_excerpt"],
                    "confidence": row["original"]["confidence"],
                }
                for row in self.state
            ],
        }
        self.reviewed_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        messagebox.showinfo("Saved", f"Saved review to:\n{self.reviewed_path}")

    def save_and_apply(self) -> None:
        self.save_review()
        invalid = []
        mod = fundi_mod()
        for row in self.state:
            if row["decision"] not in {"confirm", "error"}:
                continue
            if not mod.TRIPLE_PATTERN.search(row["corrected_triple"]):
                invalid.append(row["id"])
        if invalid:
            messagebox.showerror("Invalid Triple", f"These items have invalid corrected triples: {invalid}")
            return
        try:
            results = mod.apply_review()
        except Exception as exc:
            messagebox.showerror("Apply Failed", str(exc))
            return
        summary = "\n".join(f"Item {idx + 1}: {result['status']}" for idx, result in enumerate(results))
        messagebox.showinfo("Apply Finished", summary or "No changes were applied.")

    def copy_html_path(self) -> None:
        path = str(Path(fundi_mod().review_html_path()).resolve())
        self.winfo_toplevel().clipboard_clear()
        self.winfo_toplevel().clipboard_append(path)
        messagebox.showinfo("Copied", f"Copied HTML path:\n{path}")


class UnifiedReviewCenter(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Unified Review Center")
        self.geometry("1500x940")
        self.configure(bg=BG)
        self.notebook: ttk.Notebook | None = None
        self.tab_frames: dict[str, tk.Frame] = {}
        self._loaded_tabs: dict[str, tk.Frame] = {}
        self._configure_style()
        self._build_ui()

    def _configure_style(self) -> None:
        style = ttk.Style(self)
        if "clam" in style.theme_names():
            style.theme_use("clam")
        style.configure("App.TFrame", background=BG)
        style.configure("Panel.TFrame", background=PANEL)
        style.configure("Card.TFrame", background=CARD)
        style.configure("Title.TLabel", background=BG, foreground=TEXT, font=(FONT_UI, 20, "bold"))
        style.configure("Sub.TLabel", background=BG, foreground=MUTED, font=(FONT_UI, 10))
        style.configure("Section.TLabel", background=PANEL, foreground=TEXT, font=(FONT_UI, 10, "bold"))
        style.configure(
            "Primary.TButton",
            background=ACCENT,
            foreground="#ffffff",
            borderwidth=0,
            focuscolor=BG,
            padding=(12, 8),
        )
        style.map("Primary.TButton", background=[("active", "#1d4ed8")])
        style.configure(
            "Secondary.TButton",
            background=PANEL_2,
            foreground=TEXT,
            bordercolor=BORDER,
            lightcolor=PANEL_2,
            darkcolor=PANEL_2,
            padding=(12, 8),
        )
        style.map("Secondary.TButton", background=[("active", "#eef2f7")])
        style.configure("Accent.TButton", background="#eff6ff", foreground=ACCENT, borderwidth=0, padding=(12, 8))
        style.map("Accent.TButton", background=[("active", "#dbeafe")])
        style.configure("Danger.TButton", background="#fff1f2", foreground=ERROR, borderwidth=0, padding=(12, 8))
        style.map("Danger.TButton", background=[("active", "#ffe4e6")])
        style.configure(
            "App.TCombobox",
            fieldbackground=PANEL_2,
            background=PANEL_2,
            foreground=TEXT,
            arrowcolor=ACCENT,
            bordercolor=BORDER,
            lightcolor=PANEL_2,
            darkcolor=PANEL_2,
        )
        style.configure("Center.TNotebook", background=BG, borderwidth=0)
        style.configure("Center.TNotebook.Tab", padding=(16, 10), font=(FONT_UI, 10, "bold"))

    def _build_ui(self) -> None:
        top = ttk.Frame(self, style="App.TFrame", padding=(18, 18, 18, 8))
        top.pack(fill="x")
        ttk.Label(top, text="Unified Review Center", style="Title.TLabel").pack(anchor="w")
        ttk.Label(
            top,
            text="Literature search review and FundiGraph triplet confirmation in one window.",
            style="Sub.TLabel",
        ).pack(anchor="w", pady=(4, 4))

        self.notebook = ttk.Notebook(self, style="Center.TNotebook")
        self.notebook.pack(fill="both", expand=True, padx=18, pady=(0, 18))

        self.tab_frames["literature"] = tk.Frame(self.notebook, bg=BG)
        self.tab_frames["fundi"] = tk.Frame(self.notebook, bg=BG)

        self.notebook.add(self.tab_frames["literature"], text="Literature Search")
        self.notebook.add(self.tab_frames["fundi"], text="FundiGraph Review")
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)
        self._ensure_tab_loaded("literature")

    def _on_tab_changed(self, _event=None) -> None:
        if self.notebook is None:
            return
        current = self.notebook.select()
        if current == str(self.tab_frames["literature"]):
            self._ensure_tab_loaded("literature")
        elif current == str(self.tab_frames["fundi"]):
            self._ensure_tab_loaded("fundi")

    def _ensure_tab_loaded(self, key: str) -> None:
        if key in self._loaded_tabs:
            return
        parent = self.tab_frames[key]
        frame = LiteratureFrame(parent) if key == "literature" else FundiGraphFrame(parent)
        frame.pack(fill="both", expand=True)
        self._loaded_tabs[key] = frame


def main() -> None:
    app = UnifiedReviewCenter()
    app.mainloop()


if __name__ == "__main__":
    main()
