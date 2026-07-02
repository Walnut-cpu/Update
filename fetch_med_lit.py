from __future__ import annotations

import argparse
import json
import re
import time
from collections import OrderedDict
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import quote
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
DISEASE_FILE = BASE_DIR / "Disease_list.xlsx"
DOI_FILE = BASE_DIR / "DOI_list.xlsx"
MARKDOWN_DIR = BASE_DIR / "Markdown"
SEARCH_STORE_FILE = BASE_DIR / "search_strategy_store.json"
SEARCH_REVIEW_FILE = BASE_DIR / "search_strategy_review.json"
SEARCH_REVIEW_MARKDOWN = BASE_DIR / "search_strategy_review.md"
RUN_STATE_FILE = BASE_DIR / "literature_run_state.json"
SINCE_DATE = date(2026, 6, 5)
INITIAL_SEARCH_START_DATE = SINCE_DATE + timedelta(days=1)


@dataclass
class Record:
    disease: str
    title: str
    publication_date: str
    work_type: str
    source: str
    journal_or_venue: str
    authors: str
    doi: str
    identifier: str
    landing_page: str
    pdf_url: str
    abstract: str
    license: str
    is_open_access: str
    retrieved_from: str


RESULT_HEADERS = [
    "Disease",
    "Title",
    "Publication Date",
    "Recorded On",
    "Type",
    "Source Database",
    "Journal/Venue",
    "Authors",
    "DOI",
    "Identifier",
    "Open Access",
    "License/OA Status",
    "Landing Page",
    "PDF URL",
    "Retrieved-from Record",
    "Abstract File",
]


def today_iso() -> str:
    return datetime.now().date().isoformat()


def today_stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def read_diseases(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    values: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    text = clean_text(value)
                    if text:
                        values.append(text)
    return list(OrderedDict.fromkeys(values))


def clean_text(text: str | None) -> str:
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()


def slugify(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_") or "disease"


def normalize_disease_name(name: str) -> str:
    normalized = clean_text(name)
    normalized = normalized.replace("’", "'")
    return normalized


def abbreviation_from_name(name: str) -> str:
    stopwords = {"and", "of", "the", "with", "without", "to", "in", "for", "related"}
    parts = re.findall(r"[A-Za-z]+", name)
    letters = [part[0].upper() for part in parts if part.lower() not in stopwords]
    if 3 <= len(letters) <= 6:
        return "".join(letters)
    return ""


def unique_terms(terms: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for term in terms:
        cleaned = clean_text(term)
        key = cleaned.lower()
        if not cleaned or key in seen:
            continue
        seen.add(key)
        ordered.append(cleaned)
    return ordered


def build_europe_pmc_clause(terms: Iterable[str]) -> str:
    return " OR ".join(format_europe_pmc_term(term) for term in unique_terms(terms))


def build_stable_europe_pmc_query(terms: Iterable[str]) -> str:
    clause = build_europe_pmc_clause(terms)
    return f"({clause}) AND OPEN_ACCESS:y AND HAS_ABSTRACT:y"


def build_runtime_europe_pmc_query(terms: Iterable[str], start_date: date, end_date: date) -> str:
    clause = build_europe_pmc_clause(terms)
    return (
        f"({clause}) AND FIRST_PDATE:[{start_date.isoformat()} TO {end_date.isoformat()}] "
        "AND OPEN_ACCESS:y AND HAS_ABSTRACT:y"
    )


def generate_search_terms(disease: str) -> dict[str, Any]:
    canonical = normalize_disease_name(disease)
    terms: list[str] = [canonical]

    no_hyphen = canonical.replace("-", " ")
    if no_hyphen.lower() != canonical.lower():
        terms.append(no_hyphen)

    abbreviation = abbreviation_from_name(canonical)
    if abbreviation:
        terms.append(abbreviation)

    disease_lower = canonical.lower()
    notes = [
        "Only disease-name-centered terms are allowed.",
        "Do not expand to examination, imaging, biomarker, or treatment technique terms.",
        "Subtypes are allowed only when the subtype name still explicitly contains the disease identity.",
    ]

    if disease_lower == "age-related macular degeneration":
        terms.extend(
            [
                "age related macular degeneration",
                "AMD",
                "neovascular age-related macular degeneration",
                "nAMD",
                "dry AMD",
                "wet AMD",
                "non-neovascular age-related macular degeneration",
            ]
        )
        notes.append("Geographic atrophy is not included unless explicitly labeled as AMD-related in a later manual edit.")

    if disease_lower == "bergmeister papilla":
        notes.append("No broader ophthalmology expansions were added automatically.")

    terms = unique_terms(terms)
    return {
        "canonical_name": canonical,
        "canonical_key": slugify(canonical),
        "terms": terms,
        "notes": notes,
        "queries": {
            "openalex_terms": terms,
            "europe_pmc_query": build_stable_europe_pmc_query(terms),
        },
    }


def format_europe_pmc_term(term: str) -> str:
    if re.fullmatch(r"[A-Za-z0-9-]+", term):
        return term
    return f'"{term}"'


def load_json_file(path: Path, default: dict[str, Any]) -> dict[str, Any]:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def save_json_file(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_run_state() -> dict[str, Any]:
    default = {
        "schema_version": 1,
        "last_successful_run_on": "",
        "last_window_start": "",
        "last_window_end": "",
        "updated_at": "",
    }
    state = load_json_file(RUN_STATE_FILE, default)
    for key, value in default.items():
        state.setdefault(key, value)
    return state


def parse_iso_date(value: str | None) -> date | None:
    try:
        return date.fromisoformat(clean_text(value))
    except Exception:
        return None


def determine_search_window(run_state: dict[str, Any]) -> tuple[date, date]:
    end_date = datetime.now().date()
    last_successful_run = parse_iso_date(run_state.get("last_successful_run_on"))
    start_date = last_successful_run + timedelta(days=1) if last_successful_run else INITIAL_SEARCH_START_DATE
    if start_date > end_date:
        start_date = end_date
    return start_date, end_date


def load_store() -> dict[str, Any]:
    default = {"schema_version": 1, "generated_on": today_iso(), "strategies": {}}
    store = load_json_file(SEARCH_STORE_FILE, default)
    store.setdefault("schema_version", 1)
    store.setdefault("generated_on", today_iso())
    store.setdefault("strategies", {})
    return store


def load_review_payload() -> dict[str, Any]:
    default = {"schema_version": 1, "generated_on": today_iso(), "review_items": []}
    payload = load_json_file(SEARCH_REVIEW_FILE, default)
    payload.setdefault("schema_version", 1)
    payload.setdefault("generated_on", today_iso())
    payload.setdefault("review_items", [])
    return payload


def safe_unlink(path: Path) -> None:
    if path.exists():
        path.unlink()


def import_reviewed_strategies(store: dict[str, Any], review_payload: dict[str, Any]) -> tuple[int, int]:
    approved = 0
    rejected = 0
    strategies = store["strategies"]
    for item in review_payload.get("review_items", []):
        key = item.get("canonical_key") or slugify(item.get("disease_name", ""))
        status = (item.get("status") or "pending").lower()
        if not key:
            continue
        if status == "approved":
            item["status"] = "approved"
            item["approved_on"] = today_iso()
            strategies[key] = item
            approved += 1
        elif status == "rejected":
            rejected += 1
    if approved:
        store["generated_on"] = today_iso()
    return approved, rejected


def normalize_strategy_queries(item: dict[str, Any]) -> dict[str, Any]:
    generated = generate_search_terms(item.get("disease_name", ""))
    queries = item.setdefault("queries", {})
    terms = unique_terms(queries.get("openalex_terms") or generated["queries"]["openalex_terms"])
    queries["openalex_terms"] = terms
    queries["europe_pmc_query"] = build_stable_europe_pmc_query(terms)
    item["queries"] = queries
    item["time_window_mode"] = "last_successful_run_to_today"
    return item


def ensure_strategy_candidates(diseases: list[str], store: dict[str, Any]) -> list[dict[str, Any]]:
    strategies = store["strategies"]
    auto_generated_items: list[dict[str, Any]] = []
    for disease in diseases:
        generated = generate_search_terms(disease)
        key = generated["canonical_key"]
        existing = strategies.get(key)
        if existing and (existing.get("status") or "").lower() == "approved":
            normalize_strategy_queries(existing)
            continue
        item = {
            "disease_name": generated["canonical_name"],
            "canonical_key": key,
            "status": "approved",
            "generated_on": today_iso(),
            "approved_on": today_iso(),
            "notes": generated["notes"],
            "queries": generated["queries"],
            "review_comment": "Auto-approved generated strategy.",
            "auto_generated": True,
        }
        normalize_strategy_queries(item)
        strategies[key] = item
        auto_generated_items.append(item)
    return auto_generated_items


def write_review_files(review_items: list[dict[str, Any]]) -> None:
    payload = {
        "schema_version": 1,
        "generated_on": today_iso(),
        "instructions": [
            "Review each pending disease strategy.",
            "Change status from pending to approved or rejected in search_strategy_review.json.",
            "You may edit the query terms before approval.",
            "After approval, rerun fetch_med_lit.py and the approved strategy will move into the local store automatically.",
        ],
        "review_items": review_items,
    }
    save_json_file(SEARCH_REVIEW_FILE, payload)

    lines = [
        "# Search Strategy Review",
        "",
        "Review the candidate search strategies below.",
        "",
        "How to approve:",
        "- Open `search_strategy_review.json`.",
        "- Change `status` from `pending` to `approved` or `rejected`.",
        "- Optionally edit the query terms before approval.",
        "- Rerun `fetch_med_lit.py` after saving your review.",
        "",
    ]
    for index, item in enumerate(review_items, start=1):
        lines.extend(
            [
                f"## {index}. {item['disease_name']}",
                "",
                f"- Canonical key: `{item['canonical_key']}`",
                f"- Status: `{item['status']}`",
                "- Notes:",
            ]
        )
        for note in item["notes"]:
            lines.append(f"  - {note}")
        lines.extend(
            [
                "- OpenAlex terms:",
                f"  - `{', '.join(item['queries']['openalex_terms'])}`",
                "- Europe PMC query:",
                f"  - `{item['queries']['europe_pmc_query']}`",
                "",
            ]
        )
    SEARCH_REVIEW_MARKDOWN.write_text("\n".join(lines), encoding="utf-8")


def fetch_json(url: str) -> dict[str, Any]:
    req = Request(
        url,
        headers={
            "User-Agent": "CodexMedicalLiteratureFetcher/1.0 (mailto:example@example.com)",
            "Accept": "application/json",
        },
    )
    with urlopen(req, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def normalize_date(value: str | None) -> str:
    if not value:
        return ""
    return value[:10]


def iso_in_window(value: str, start_date: date, end_date: date) -> bool:
    try:
        published = date.fromisoformat(value[:10])
    except Exception:
        return False
    return start_date <= published <= end_date


def invert_abstract(index: dict[str, list[int]] | None) -> str:
    if not index:
        return ""
    positions: list[tuple[int, str]] = []
    for word, pos_list in index.items():
        for pos in pos_list:
            positions.append((int(pos), word))
    positions.sort(key=lambda item: item[0])
    return clean_text(" ".join(word for _, word in positions))


def authors_from_openalex(authorships: list[dict[str, Any]] | None) -> str:
    if not authorships:
        return ""
    names = []
    for authorship in authorships[:10]:
        author = authorship.get("author") or {}
        name = author.get("display_name")
        if name:
            names.append(name)
    return "; ".join(names)


def venue_from_openalex(item: dict[str, Any]) -> str:
    primary_location = item.get("primary_location") or {}
    source = primary_location.get("source") or {}
    return source.get("display_name") or ""


def extract_identifier(item: dict[str, Any]) -> str:
    ids = item.get("ids") or {}
    return ids.get("pmid") or ids.get("pmcid") or ids.get("openalex") or item.get("id") or ""


def landing_page(item: dict[str, Any]) -> tuple[str, str]:
    primary_location = item.get("primary_location") or {}
    landing = primary_location.get("landing_page_url") or item.get("id") or ""
    pdf = primary_location.get("pdf_url") or ""
    return landing, pdf


def query_openalex(disease: str, terms: list[str], start_date: date, end_date: date) -> list[Record]:
    records: list[Record] = []
    seen: set[str] = set()
    for term in terms:
        cursor = "*"
        while True:
            url = (
                "https://api.openalex.org/works?"
                f"search={quote(term)}"
                f"&filter=from_publication_date:{start_date.isoformat()},to_publication_date:{end_date.isoformat()},is_oa:true,has_abstract:true"
                "&per-page=200"
                f"&cursor={quote(cursor)}"
            )
            payload = fetch_json(url)
            results = payload.get("results") or []
            for item in results:
                pub_date = normalize_date(item.get("publication_date"))
                if not iso_in_window(pub_date, start_date, end_date):
                    continue
                title = clean_text(item.get("display_name"))
                if not title:
                    continue
                key = item.get("id") or f"{title}|{pub_date}"
                if key in seen:
                    continue
                seen.add(key)
                abstract = invert_abstract(item.get("abstract_inverted_index"))
                if not abstract:
                    continue
                doi = item.get("doi") or ""
                work_type = item.get("type") or ""
                landing, pdf = landing_page(item)
                oa = item.get("open_access") or {}
                records.append(
                    Record(
                        disease=disease,
                        title=title,
                        publication_date=pub_date,
                        work_type=work_type,
                        source="OpenAlex",
                        journal_or_venue=venue_from_openalex(item),
                        authors=authors_from_openalex(item.get("authorships")),
                        doi=doi.replace("https://doi.org/", ""),
                        identifier=extract_identifier(item),
                        landing_page=landing,
                        pdf_url=pdf,
                        abstract=abstract,
                        license=oa.get("oa_status") or "",
                        is_open_access="Yes",
                        retrieved_from=f"{item.get('id') or ''} | term={term}",
                    )
                )
            meta = payload.get("meta") or {}
            next_cursor = meta.get("next_cursor")
            if not next_cursor or next_cursor == cursor or not results:
                break
            cursor = next_cursor
            time.sleep(0.2)
    return records


def authors_from_epmc(author_string: str | None) -> str:
    return clean_text(author_string)


def doi_from_epmc(item: dict[str, Any]) -> str:
    value = item.get("doi") or ""
    return value.replace("https://doi.org/", "")


def identifier_from_epmc(item: dict[str, Any]) -> str:
    for key in ("pmid", "pmcid", "id"):
        value = item.get(key)
        if value:
            return str(value)
    return ""


def link_from_epmc(item: dict[str, Any]) -> tuple[str, str]:
    pmcid = item.get("pmcid")
    if pmcid:
        return f"https://pmc.ncbi.nlm.nih.gov/articles/{pmcid}/", ""
    doi = doi_from_epmc(item)
    if doi:
        return f"https://doi.org/{doi}", ""
    pmid = item.get("pmid")
    if pmid:
        return f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/", ""
    return "", ""


def query_europe_pmc(disease: str, query_string: str, start_date: date, end_date: date) -> list[Record]:
    records: list[Record] = []
    seen: set[str] = set()
    page = 1
    while True:
        url = (
            "https://www.ebi.ac.uk/europepmc/webservices/rest/search?"
            f"query={quote(query_string)}"
            "&format=json"
            "&pageSize=1000"
            f"&page={page}"
            "&resultType=core"
        )
        payload = fetch_json(url)
        result_list = (((payload or {}).get("resultList")) or {}).get("result") or []
        for item in result_list:
            pub_date = normalize_date(item.get("firstPublicationDate") or item.get("firstIndexDate"))
            if not iso_in_window(pub_date, start_date, end_date):
                continue
            title = clean_text(item.get("title"))
            if not title:
                continue
            key = f"{title}|{pub_date}|{item.get('source')}|{item.get('id')}"
            if key in seen:
                continue
            seen.add(key)
            abstract = clean_text(item.get("abstractText"))
            if not abstract:
                continue
            landing, pdf = link_from_epmc(item)
            records.append(
                Record(
                    disease=disease,
                    title=title,
                    publication_date=pub_date,
                    work_type=clean_text(item.get("pubType")) or clean_text(item.get("resultType")),
                    source="Europe PMC",
                    journal_or_venue=clean_text(item.get("journalTitle")),
                    authors=authors_from_epmc(item.get("authorString")),
                    doi=doi_from_epmc(item),
                    identifier=identifier_from_epmc(item),
                    landing_page=landing,
                    pdf_url=pdf,
                    abstract=abstract,
                    license=clean_text(item.get("license")),
                    is_open_access="Yes",
                    retrieved_from=f"{item.get('source', '')}:{item.get('id', '')}",
                )
            )
        hit_count = int((((payload or {}).get("hitCount")) or 0))
        if page * 1000 >= hit_count or not result_list:
            break
        page += 1
        time.sleep(0.2)
    return records


def dedupe_records(records: Iterable[Record]) -> list[Record]:
    deduped: OrderedDict[str, Record] = OrderedDict()
    for record in sorted(records, key=lambda r: (r.disease.lower(), r.publication_date, r.title.lower())):
        key = (record.disease.lower(), record.doi.lower() if record.doi else "", record.identifier.lower(), record.title.lower())
        deduped[str(key)] = record
    return list(deduped.values())


def record_unique_id(record: Record) -> str:
    if record.doi:
        return f"doi:{record.doi.lower()}"
    if record.identifier:
        return f"id:{record.identifier.lower()}"
    return f"title:{record.title.lower()}"


def record_key(record: Record) -> str:
    return str(
        (
            record.disease.lower(),
            record.doi.lower() if record.doi else "",
            record.identifier.lower(),
            record.title.lower(),
        )
    )


def sanitize_filename(value: str) -> str:
    value = re.sub(r"[<>:\"/\\\\|?*]+", "_", value).strip()
    value = re.sub(r"\s+", "_", value)
    return value[:140] or "record"


def daily_markdown_filename() -> str:
    return f"{today_stamp()}.md"


def load_existing_results() -> tuple[list[Record], dict[str, str], dict[str, str]]:
    if not DOI_FILE.exists():
        return [], {}, {}
    try:
        wb = load_workbook(DOI_FILE, read_only=True, data_only=True)
    except Exception:
        return [], {}, {}
    if "Results" not in wb.sheetnames:
        return [], {}, {}

    ws = wb["Results"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {}, {}

    header = [clean_text(str(cell)) if cell is not None else "" for cell in rows[0]]
    index = {name: idx for idx, name in enumerate(header)}
    required_headers = [header for header in RESULT_HEADERS if header != "Recorded On"]
    if not set(required_headers).issubset(index.keys()):
        return [], {}, {}

    records: list[Record] = []
    abstract_files: dict[str, str] = {}
    recorded_on_map: dict[str, str] = {}
    for row in rows[1:]:
        if not row or not any(value is not None and str(value).strip() for value in row):
            continue
        record = Record(
            disease=clean_text(row[index["Disease"]]) if row[index["Disease"]] else "",
            title=clean_text(row[index["Title"]]) if row[index["Title"]] else "",
            publication_date=clean_text(row[index["Publication Date"]]) if row[index["Publication Date"]] else "",
            work_type=clean_text(row[index["Type"]]) if row[index["Type"]] else "",
            source=clean_text(row[index["Source Database"]]) if row[index["Source Database"]] else "",
            journal_or_venue=clean_text(row[index["Journal/Venue"]]) if row[index["Journal/Venue"]] else "",
            authors=clean_text(row[index["Authors"]]) if row[index["Authors"]] else "",
            doi=clean_text(row[index["DOI"]]) if row[index["DOI"]] else "",
            identifier=clean_text(row[index["Identifier"]]) if row[index["Identifier"]] else "",
            landing_page=clean_text(row[index["Landing Page"]]) if row[index["Landing Page"]] else "",
            pdf_url=clean_text(row[index["PDF URL"]]) if row[index["PDF URL"]] else "",
            abstract="",
            license=clean_text(row[index["License/OA Status"]]) if row[index["License/OA Status"]] else "",
            is_open_access=clean_text(row[index["Open Access"]]) if row[index["Open Access"]] else "",
            retrieved_from=clean_text(row[index["Retrieved-from Record"]]) if row[index["Retrieved-from Record"]] else "",
        )
        key = record_key(record)
        records.append(record)
        if "Recorded On" in index:
            recorded_on = clean_text(row[index["Recorded On"]]) if row[index["Recorded On"]] else ""
            if recorded_on:
                recorded_on_map[key] = recorded_on
        abstract_name = clean_text(row[index["Abstract File"]]) if row[index["Abstract File"]] else ""
        if abstract_name:
            abstract_files[key] = abstract_name
    return records, abstract_files, recorded_on_map


def next_markdown_index(existing_names: Iterable[str]) -> int:
    max_index = 0
    for name in existing_names:
        match = re.match(r"^(\d{3,})_", name)
        if match:
            max_index = max(max_index, int(match.group(1)))
    return max_index + 1


def write_markdown(records: list[Record], abstract_file_map: dict[str, str]) -> dict[str, str]:
    if not records:
        return abstract_file_map
    MARKDOWN_DIR.mkdir(parents=True, exist_ok=True)
    file_name = daily_markdown_filename()
    path = MARKDOWN_DIR / file_name
    header = [
        f"# New Literature Retrieved On {today_stamp()}",
        "",
        f"Generated on: {today_iso()}",
        "",
    ]
    existing_text = path.read_text(encoding="utf-8") if path.exists() else ""
    chunks: list[str] = []
    if not existing_text:
        chunks.append("\n".join(header))
    for record in records:
        key = record_key(record)
        abstract_file_map[key] = file_name
        chunks.append(
            "\n".join(
                [
                    f"## {record.title}",
                    "",
                    f"- Disease: {record.disease}",
                    f"- Publication date: {record.publication_date}",
                    f"- Type: {record.work_type}",
                    f"- Source database: {record.source}",
                    f"- Journal/Venue: {record.journal_or_venue}",
                    f"- Authors: {record.authors}",
                    f"- DOI: {record.doi}",
                    f"- Identifier: {record.identifier}",
                    f"- Open access: {record.is_open_access}",
                    f"- License/OA status: {record.license}",
                    f"- Landing page: {record.landing_page}",
                    f"- PDF URL: {record.pdf_url}",
                    f"- Retrieved-from record: {record.retrieved_from}",
                    "",
                    "### Abstract",
                    "",
                    record.abstract,
                    "",
                ]
            )
        )
    append_text = "\n".join(chunks).strip()
    if append_text:
        final_text = existing_text.rstrip() + "\n\n" + append_text if existing_text.strip() else append_text
        path.write_text(final_text.strip() + "\n", encoding="utf-8")
    return abstract_file_map


def write_excel(
    records: list[Record],
    store: dict[str, Any],
    generated_items: list[dict[str, Any]],
    imported_review_count: int,
    abstract_file_map: dict[str, str],
    recorded_on_map: dict[str, str],
    search_window_start: date,
    search_window_end: date,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(RESULT_HEADERS)
    for record in records:
        md_name = abstract_file_map.get(record_key(record), "")
        ws.append(
            [
                record.disease,
                record.title,
                record.publication_date,
                recorded_on_map.get(record_key(record), today_iso()),
                record.work_type,
                record.source,
                record.journal_or_venue,
                record.authors,
                record.doi,
                record.identifier,
                record.is_open_access,
                record.license,
                record.landing_page,
                record.pdf_url,
                record.retrieved_from,
                md_name,
            ]
        )

    meta = wb.create_sheet("Metadata")
    meta.append(["Field", "Value"])
    meta.append(["Generated On", today_iso()])
    meta.append(["Initial Search Bootstrap Date", INITIAL_SEARCH_START_DATE.isoformat()])
    meta.append(["Search Window Start", search_window_start.isoformat()])
    meta.append(["Search Window End", search_window_end.isoformat()])
    meta.append(["Approved Strategies Imported This Run", imported_review_count])
    meta.append(["Auto-generated Strategy Count", len(generated_items)])
    meta.append(["Pending Strategy Count", 0])
    meta.append(["Search Strategy Store", SEARCH_STORE_FILE.name])
    meta.append(["Run State File", RUN_STATE_FILE.name])
    meta.append(["Review File", "not used for new strategy generation"])
    meta.append(["New Literature Markdown File", daily_markdown_filename()])

    strategy_sheet = wb.create_sheet("Search_Strategy")
    strategy_sheet.append(["Disease", "Canonical Key", "Status", "Time Window Mode", "OpenAlex Terms", "Europe PMC Query"])
    approved_items = sorted(store["strategies"].values(), key=lambda item: item.get("disease_name", "").lower())
    for item in approved_items:
        strategy_sheet.append(
            [
                item.get("disease_name"),
                item.get("canonical_key"),
                item.get("status"),
                item.get("time_window_mode", ""),
                ", ".join(item.get("queries", {}).get("openalex_terms", [])),
                item.get("queries", {}).get("europe_pmc_query", ""),
            ]
        )

    generated_sheet = wb.create_sheet("Auto_Generated_Strategy")
    generated_sheet.append(["Disease", "Canonical Key", "Status", "Review Comment"])
    for item in generated_items:
        generated_sheet.append(
            [
                item.get("disease_name"),
                item.get("canonical_key"),
                item.get("status"),
                item.get("review_comment", ""),
            ]
        )

    wb.save(DOI_FILE)


def sync_search_strategy_state(diseases: list[str], store: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]], int]:
    review_payload = load_review_payload()
    imported_review_count, _ = import_reviewed_strategies(store, review_payload)
    for item in store["strategies"].values():
        normalize_strategy_queries(item)

    generated_items = ensure_strategy_candidates(diseases, store)
    safe_unlink(SEARCH_REVIEW_FILE)
    safe_unlink(SEARCH_REVIEW_MARKDOWN)

    save_json_file(SEARCH_STORE_FILE, store)
    return store, generated_items, imported_review_count


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--strategy-only",
        action="store_true",
        help="Only sync local search strategy files and skip all network retrieval.",
    )
    args = parser.parse_args()

    diseases = read_diseases(DISEASE_FILE)
    store = load_store()
    run_state = load_run_state()
    search_window_start, search_window_end = determine_search_window(run_state)
    store, generated_items, imported_review_count = sync_search_strategy_state(diseases, store)

    if args.strategy_only:
        print(
            json.dumps(
                {
                    "mode": "strategy_only",
                    "diseases": diseases,
                    "approved_strategy_count": len(store.get("strategies", {})),
                    "auto_generated_strategy_count": len(generated_items),
                    "pending_strategy_count": 0,
                    "search_window_start": search_window_start.isoformat(),
                    "search_window_end": search_window_end.isoformat(),
                    "review_file": "",
                    "store_file": SEARCH_STORE_FILE.name,
                    "network_retrieval_skipped": True,
                },
                ensure_ascii=False,
            )
        )
        return

    all_records: list[Record] = []
    approved_count = 0
    for disease in diseases:
        key = slugify(normalize_disease_name(disease))
        strategy = store["strategies"].get(key)
        if not strategy or (strategy.get("status") or "").lower() != "approved":
            continue
        approved_count += 1
        queries = strategy.get("queries", {})
        openalex_terms = unique_terms(queries.get("openalex_terms", []))
        europe_pmc_query = build_runtime_europe_pmc_query(openalex_terms, search_window_start, search_window_end)
        all_records.extend(query_openalex(disease, openalex_terms, search_window_start, search_window_end))
        all_records.extend(query_europe_pmc(disease, europe_pmc_query, search_window_start, search_window_end))

    existing_records, abstract_file_map, recorded_on_map = load_existing_results()
    existing_unique_ids = {record_unique_id(record) for record in existing_records}
    new_unique_records = [record for record in dedupe_records(all_records) if record_unique_id(record) not in existing_unique_ids]
    merged_records = dedupe_records([*existing_records, *new_unique_records])
    if approved_count:
        for record in new_unique_records:
            recorded_on_map[record_key(record)] = today_iso()
        abstract_file_map = write_markdown(new_unique_records, abstract_file_map)
        write_excel(
            merged_records,
            store,
            generated_items,
            imported_review_count,
            abstract_file_map,
            recorded_on_map,
            search_window_start,
            search_window_end,
        )
        run_state["last_successful_run_on"] = today_iso()
        run_state["last_window_start"] = search_window_start.isoformat()
        run_state["last_window_end"] = search_window_end.isoformat()
        run_state["updated_at"] = datetime.now().isoformat(timespec="seconds")
        save_json_file(RUN_STATE_FILE, run_state)
    print(
        json.dumps(
            {
                "diseases": diseases,
                "approved_strategy_count": approved_count,
                "auto_generated_strategy_count": len(generated_items),
                "pending_strategy_count": 0,
                "search_window_start": search_window_start.isoformat(),
                "search_window_end": search_window_end.isoformat(),
                "new_record_count": len(new_unique_records),
                "total_record_count": len(merged_records),
                "review_file": "",
                "store_file": SEARCH_STORE_FILE.name,
                "run_state_file": RUN_STATE_FILE.name,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
