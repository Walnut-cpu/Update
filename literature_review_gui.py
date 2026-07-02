from __future__ import annotations

import json
import tkinter as tk
from pathlib import Path
from tkinter import messagebox, ttk

from openpyxl import load_workbook

from fetch_med_lit import (
    BASE_DIR,
    DISEASE_FILE,
    DOI_FILE,
    SEARCH_REVIEW_FILE,
    SEARCH_REVIEW_MARKDOWN,
    SEARCH_STORE_FILE,
    ensure_strategy_candidates,
    generate_search_terms,
    load_json_file,
    load_store,
    read_diseases,
    save_json_file,
    today_iso,
)


BG = "#f5f7fb"
PANEL = "#ffffff"
CARD = "#ffffff"
PANEL_2 = "#f8fafc"
ACCENT = "#2563eb"
TEXT = "#0f172a"
MUTED = "#64748b"
BORDER = "#dbe3f0"
SUCCESS = "#15803d"
WARN = "#b45309"
ERROR = "#dc2626"
FONT_UI = "Segoe UI"
FONT_MONO = "Consolas"


class LiteratureReviewApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Literature Search Desk")
        self.root.geometry("1460x920")
        self.root.configure(bg=BG)

        self.diseases: list[str] = []
        self.store: dict = {}
        self.review_payload: dict = {}
        self.results_rows: list[dict[str, str]] = []
        self.no_result_rows: dict[str, dict[str, str]] = {}
        self.disease_items: dict[str, dict] = {}
        self.current_key: str | None = None

        self.status_var = tk.StringVar()
        self.summary_var = tk.StringVar()
        self.filter_var = tk.StringVar(value="all")
        self.strategy_status_var = tk.StringVar(value="pending")
        self.search_var = tk.StringVar()

        self.term_text: tk.Text | None = None
        self.query_text: tk.Text | None = None
        self.notes_text: tk.Text | None = None
        self.comment_text: tk.Text | None = None
        self.overview_text: tk.Text | None = None
        self.result_tree: ttk.Treeview | None = None
        self.no_result_tree: ttk.Treeview | None = None
        self.disease_list: tk.Listbox | None = None

        self._configure_style()
        self._build_ui()
        self.reload_all()

    def _configure_style(self) -> None:
        style = ttk.Style(self.root)
        if "clam" in style.theme_names():
            style.theme_use("clam")
        style.configure("App.TFrame", background=BG)
        style.configure("Panel.TFrame", background=PANEL)
        style.configure("Card.TFrame", background=CARD)
        style.configure("Title.TLabel", background=BG, foreground=TEXT, font=(FONT_UI, 20, "bold"))
        style.configure("Sub.TLabel", background=BG, foreground=MUTED, font=(FONT_UI, 10))
        style.configure("Section.TLabel", background=PANEL, foreground=TEXT, font=(FONT_UI, 10, "bold"))
        style.configure(
            "Primary.TButton",
            background=ACCENT,
            foreground="#ffffff",
            borderwidth=0,
            padding=(12, 8),
        )
        style.map("Primary.TButton", background=[("active", "#1d4ed8")])
        style.configure(
            "Secondary.TButton",
            background=PANEL_2,
            foreground=TEXT,
            bordercolor=BORDER,
            lightcolor=PANEL_2,
            darkcolor=PANEL_2,
            padding=(12, 8),
        )
        style.map("Secondary.TButton", background=[("active", "#eef2f7")])
        style.configure(
            "Accent.TButton",
            background="#eff6ff",
            foreground=ACCENT,
            borderwidth=0,
            padding=(12, 8),
        )
        style.map("Accent.TButton", background=[("active", "#dbeafe")])
        style.configure("App.TNotebook", background=BG, borderwidth=0)
        style.configure("App.TNotebook.Tab", padding=(12, 8), font=(FONT_UI, 10))
        style.configure("App.Treeview", rowheight=28, font=(FONT_UI, 9), background=CARD, fieldbackground=CARD)
        style.configure("App.Treeview.Heading", font=(FONT_UI, 9, "bold"))
        style.configure(
            "App.TCombobox",
            fieldbackground=PANEL_2,
            background=PANEL_2,
            foreground=TEXT,
            arrowcolor=ACCENT,
            bordercolor=BORDER,
            lightcolor=PANEL_2,
            darkcolor=PANEL_2,
        )

    def _build_ui(self) -> None:
        top = ttk.Frame(self.root, style="App.TFrame", padding=18)
        top.pack(fill="x")
        ttk.Label(top, text="Literature Search Desk", style="Title.TLabel").pack(anchor="w")
        ttk.Label(
            top,
            text="Review diseases, manage search strategies, and inspect retrieved literature from one screen.",
            style="Sub.TLabel",
        ).pack(anchor="w", pady=(4, 12))

        toolbar = ttk.Frame(top, style="App.TFrame")
        toolbar.pack(fill="x")
        ttk.Button(toolbar, text="Reload", style="Secondary.TButton", command=self.reload_all).pack(side="left")
        ttk.Button(toolbar, text="Generate Missing Strategies", style="Accent.TButton", command=self.generate_missing).pack(
            side="left", padx=(10, 0)
        )
        ttk.Button(toolbar, text="Save Current", style="Secondary.TButton", command=self.save_current_item).pack(
            side="left", padx=(10, 0)
        )
        ttk.Button(toolbar, text="Save All", style="Primary.TButton", command=self.save_all).pack(side="left", padx=(10, 0))
        ttk.Label(toolbar, text="Filter", style="Sub.TLabel").pack(side="left", padx=(18, 8))
        filter_box = ttk.Combobox(
            toolbar,
            textvariable=self.filter_var,
            values=["all", "approved", "pending", "rejected", "missing"],
            state="readonly",
            width=12,
            style="App.TCombobox",
        )
        filter_box.pack(side="left")
        filter_box.bind("<<ComboboxSelected>>", lambda _e: self.refresh_disease_list())

        search_entry = tk.Entry(toolbar, textvariable=self.search_var, bd=0, relief="flat", font=(FONT_UI, 10))
        search_entry.pack(side="left", padx=(12, 0), ipadx=80, ipady=6)
        search_entry.bind("<KeyRelease>", lambda _e: self.refresh_disease_list())

        ttk.Label(toolbar, textvariable=self.summary_var, style="Sub.TLabel").pack(side="right")

        body = ttk.Frame(self.root, style="App.TFrame", padding=(18, 0, 18, 18))
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=0)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        left = tk.Frame(body, bg=PANEL, highlightbackground=BORDER, highlightthickness=1)
        left.grid(row=0, column=0, sticky="nsw", padx=(0, 14))
        left.configure(width=330)
        left.grid_propagate(False)

        ttk.Label(left, text="Diseases", style="Section.TLabel").pack(anchor="w", padx=14, pady=(14, 8))
        self.disease_list = tk.Listbox(
            left,
            bg=PANEL,
            fg=TEXT,
            activestyle="none",
            relief="flat",
            highlightthickness=0,
            bd=0,
            font=(FONT_UI, 10),
        )
        self.disease_list.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.disease_list.bind("<<ListboxSelect>>", self.on_select_disease)

        right = ttk.Notebook(body, style="App.TNotebook")
        right.grid(row=0, column=1, sticky="nsew")

        strategy_tab = tk.Frame(right, bg=PANEL)
        strategy_tab.grid_columnconfigure(0, weight=1)
        right.add(strategy_tab, text="Strategy Review")

        results_tab = tk.Frame(right, bg=PANEL)
        results_tab.grid_columnconfigure(0, weight=1)
        results_tab.grid_rowconfigure(1, weight=1)
        right.add(results_tab, text="Literature Results")

        self._build_strategy_tab(strategy_tab)
        self._build_results_tab(results_tab)

    def _build_strategy_tab(self, parent: tk.Frame) -> None:
        header = tk.Frame(parent, bg=PANEL)
        header.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 10))
        header.grid_columnconfigure(0, weight=1)
        tk.Label(header, textvariable=self.status_var, bg=PANEL, fg=TEXT, font=(FONT_UI, 14, "bold")).grid(
            row=0, column=0, sticky="w"
        )

        form = tk.Frame(parent, bg=PANEL)
        form.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 16))
        form.grid_columnconfigure(0, weight=1)

        status_row = tk.Frame(form, bg=PANEL)
        status_row.pack(fill="x", pady=(0, 10))
        tk.Label(status_row, text="Strategy Status", bg=PANEL, fg=TEXT, font=(FONT_UI, 10, "bold")).pack(side="left")
        status_box = ttk.Combobox(
            status_row,
            textvariable=self.strategy_status_var,
            values=["pending", "approved", "rejected"],
            state="readonly",
            width=12,
            style="App.TCombobox",
        )
        status_box.pack(side="left", padx=(12, 0))

        self.overview_text = self._text_block(form, "Overview", 5, readonly=True)
        self.term_text = self._text_block(form, "OpenAlex Terms (one per line)", 8)
        self.query_text = self._text_block(form, "Europe PMC Query", 6)
        self.notes_text = self._text_block(form, "Notes (one per line)", 6)
        self.comment_text = self._text_block(form, "Review Comment", 4)

    def _build_results_tab(self, parent: tk.Frame) -> None:
        top = tk.Frame(parent, bg=PANEL)
        top.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 10))
        tk.Label(top, text="Results for selected disease", bg=PANEL, fg=TEXT, font=(FONT_UI, 10, "bold")).pack(anchor="w")

        split = tk.PanedWindow(parent, orient="vertical", bg=PANEL, sashwidth=6, bd=0, relief="flat")
        split.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 16))

        result_frame = tk.Frame(split, bg=PANEL)
        no_result_frame = tk.Frame(split, bg=PANEL)
        split.add(result_frame, stretch="always")
        split.add(no_result_frame, stretch="always")

        tk.Label(result_frame, text="Literature Records", bg=PANEL, fg=TEXT, font=(FONT_UI, 10, "bold")).pack(anchor="w")
        self.result_tree = ttk.Treeview(
            result_frame,
            columns=("date", "type", "source", "identifier", "title"),
            show="headings",
            style="App.Treeview",
            height=12,
        )
        for col, text, width in [
            ("date", "Date", 110),
            ("type", "Type", 120),
            ("source", "Source", 120),
            ("identifier", "Identifier", 180),
            ("title", "Title", 650),
        ]:
            self.result_tree.heading(col, text=text)
            self.result_tree.column(col, width=width, anchor="w")
        self.result_tree.pack(fill="both", expand=True, pady=(8, 0))

        tk.Label(no_result_frame, text="No-Result Summary", bg=PANEL, fg=TEXT, font=(FONT_UI, 10, "bold")).pack(anchor="w")
        self.no_result_tree = ttk.Treeview(
            no_result_frame,
            columns=("status", "date_filter", "checked_sources"),
            show="headings",
            style="App.Treeview",
            height=6,
        )
        for col, text, width in [
            ("status", "Status", 420),
            ("date_filter", "Date Filter", 140),
            ("checked_sources", "Checked Sources", 460),
        ]:
            self.no_result_tree.heading(col, text=text)
            self.no_result_tree.column(col, width=width, anchor="w")
        self.no_result_tree.pack(fill="both", expand=True, pady=(8, 0))

    def _text_block(self, parent: tk.Widget, title: str, height: int, readonly: bool = False) -> tk.Text:
        wrap = tk.Frame(parent, bg=PANEL)
        wrap.pack(fill="both", expand=False, pady=(0, 10))
        tk.Label(wrap, text=title, bg=PANEL, fg=TEXT, font=(FONT_UI, 10, "bold")).pack(anchor="w", pady=(0, 6))
        text = tk.Text(
            wrap,
            height=height,
            wrap="word",
            bg=CARD,
            fg=TEXT,
            insertbackground=TEXT,
            relief="flat",
            bd=0,
            padx=12,
            pady=12,
            highlightbackground=BORDER,
            highlightthickness=1,
            font=(FONT_MONO if not readonly else FONT_UI, 10),
        )
        text.pack(fill="both", expand=True)
        if readonly:
            text.configure(state="disabled")
        return text

    def reload_all(self) -> None:
        self.diseases = read_diseases(DISEASE_FILE) if DISEASE_FILE.exists() else []
        self.store = load_store()
        self.review_payload = load_json_file(
            SEARCH_REVIEW_FILE,
            {"schema_version": 1, "generated_on": today_iso(), "review_items": []},
        )
        self.results_rows, self.no_result_rows = self._load_workbook_data()
        self._build_disease_items()
        self.refresh_disease_list()
        self.status_var.set("Select a disease to review or inspect.")

    def _load_workbook_data(self) -> tuple[list[dict[str, str]], dict[str, dict[str, str]]]:
        if not DOI_FILE.exists():
            return [], {}
        wb = load_workbook(DOI_FILE, read_only=True, data_only=True)
        results_rows: list[dict[str, str]] = []
        no_result_rows: dict[str, dict[str, str]] = {}

        if "Results" in wb.sheetnames:
            ws = wb["Results"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(cell) if cell is not None else "" for cell in rows[0]]
                for row in rows[1:]:
                    if not row or not any(cell is not None for cell in row):
                        continue
                    row_map = {header[i]: ("" if row[i] is None else str(row[i])) for i in range(min(len(header), len(row)))}
                    results_rows.append(row_map)

        if "No_Result_Summary" in wb.sheetnames:
            ws = wb["No_Result_Summary"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(cell) if cell is not None else "" for cell in rows[0]]
                for row in rows[1:]:
                    if not row or not row[0]:
                        continue
                    row_map = {header[i]: ("" if row[i] is None else str(row[i])) for i in range(min(len(header), len(row)))}
                    no_result_rows[row_map["Disease"]] = row_map

        return results_rows, no_result_rows

    def _build_disease_items(self) -> None:
        self.disease_items = {}
        approved = self.store.get("strategies", {})
        pending_list = self.review_payload.get("review_items", [])
        pending_map = {item["canonical_key"]: item for item in pending_list}

        names = set(self.diseases)
        for item in approved.values():
            names.add(item.get("disease_name", ""))
        for item in pending_list:
            names.add(item.get("disease_name", ""))
        for row in self.results_rows:
            names.add(row.get("Disease", ""))
        for name in self.no_result_rows:
            names.add(name)

        for disease in sorted(name for name in names if name):
            generated = generate_search_terms(disease)
            key = generated["canonical_key"]
            strategy = approved.get(key)
            review_item = pending_map.get(key)
            if strategy:
                status = strategy.get("status", "approved")
            elif review_item:
                status = review_item.get("status", "pending")
            else:
                status = "missing"
            self.disease_items[key] = {
                "disease_name": disease,
                "canonical_key": key,
                "status": status,
                "strategy": strategy,
                "review_item": review_item,
                "generated": generated,
            }

    def refresh_disease_list(self) -> None:
        if self.disease_list is None:
            return
        self.disease_list.delete(0, tk.END)
        visible_keys: list[str] = []
        query = self.search_var.get().strip().lower()
        mode = self.filter_var.get()

        for key, item in self.disease_items.items():
            disease = item["disease_name"]
            status = item["status"]
            if query and query not in disease.lower():
                continue
            if mode != "all" and status != mode:
                continue
            visible_keys.append(key)
            self.disease_list.insert(tk.END, f"{self._status_badge(status)}  {disease}")

        self.visible_keys = visible_keys
        counts = {
            "approved": sum(1 for item in self.disease_items.values() if item["status"] == "approved"),
            "pending": sum(1 for item in self.disease_items.values() if item["status"] == "pending"),
            "rejected": sum(1 for item in self.disease_items.values() if item["status"] == "rejected"),
            "missing": sum(1 for item in self.disease_items.values() if item["status"] == "missing"),
        }
        self.summary_var.set(
            f"Approved {counts['approved']}   Pending {counts['pending']}   Missing {counts['missing']}   Rejected {counts['rejected']}"
        )

        if visible_keys:
            first = visible_keys[0]
            self.disease_list.selection_set(0)
            self.show_disease(first)
        else:
            self.current_key = None
            self.status_var.set("No disease matches the current filter.")
            self._fill_text(self.overview_text, "")
            self._fill_text(self.term_text, "")
            self._fill_text(self.query_text, "")
            self._fill_text(self.notes_text, "")
            self._fill_text(self.comment_text, "")
            self._refresh_results_view("")

    def _status_badge(self, status: str) -> str:
        mapping = {
            "approved": "APP",
            "pending": "PND",
            "rejected": "REJ",
            "missing": "NEW",
        }
        return mapping.get(status, "UNK")

    def on_select_disease(self, _event=None) -> None:
        if self.disease_list is None:
            return
        selection = self.disease_list.curselection()
        if not selection:
            return
        key = self.visible_keys[selection[0]]
        self.show_disease(key)

    def show_disease(self, key: str) -> None:
        self.current_key = key
        item = self.disease_items[key]
        disease = item["disease_name"]
        self.status_var.set(f"{disease}   {item['status'].upper()}")

        strategy = item["strategy"] or item["review_item"]
        generated = item["generated"]

        terms = (strategy or {}).get("queries", {}).get("openalex_terms", generated["queries"]["openalex_terms"])
        query = (strategy or {}).get("queries", {}).get("europe_pmc_query", generated["queries"]["europe_pmc_query"])
        notes = (strategy or {}).get("notes", generated["notes"])
        comment = (strategy or {}).get("review_comment", "")
        status = (strategy or {}).get("status", item["status"])

        overview_lines = [
            f"Disease: {disease}",
            f"Canonical key: {generated['canonical_key']}",
            f"Current status: {status}",
            f"In disease list: {'Yes' if disease in self.diseases else 'No'}",
            f"Stored strategy: {'Yes' if item['strategy'] else 'No'}",
            f"Pending review item: {'Yes' if item['review_item'] else 'No'}",
            f"Literature rows: {sum(1 for row in self.results_rows if row.get('Disease') == disease)}",
            f"No-result status: {'Yes' if disease in self.no_result_rows else 'No'}",
        ]

        self.strategy_status_var.set(status if status in {"pending", "approved", "rejected"} else "pending")
        self._fill_text(self.overview_text, "\n".join(overview_lines), readonly=True)
        self._fill_text(self.term_text, "\n".join(terms))
        self._fill_text(self.query_text, query)
        self._fill_text(self.notes_text, "\n".join(notes))
        self._fill_text(self.comment_text, comment)
        self._refresh_results_view(disease)

    def _fill_text(self, widget: tk.Text | None, value: str, readonly: bool = False) -> None:
        if widget is None:
            return
        widget.configure(state="normal")
        widget.delete("1.0", tk.END)
        widget.insert("1.0", value)
        if readonly:
            widget.configure(state="disabled")

    def _refresh_results_view(self, disease: str) -> None:
        if self.result_tree is not None:
            for item_id in self.result_tree.get_children():
                self.result_tree.delete(item_id)
            for row in self.results_rows:
                if row.get("Disease") != disease:
                    continue
                self.result_tree.insert(
                    "",
                    "end",
                    values=(
                        row.get("Publication Date", ""),
                        row.get("Type", ""),
                        row.get("Source Database", ""),
                        row.get("Identifier", "") or row.get("DOI", ""),
                        row.get("Title", ""),
                    ),
                )

        if self.no_result_tree is not None:
            for item_id in self.no_result_tree.get_children():
                self.no_result_tree.delete(item_id)
            row = self.no_result_rows.get(disease)
            if row:
                self.no_result_tree.insert(
                    "",
                    "end",
                    values=(row.get("Status", ""), row.get("Date Filter", ""), row.get("Checked Sources", "")),
                )

    def _build_payload_from_editor(self) -> dict | None:
        if not self.current_key:
            return None
        item = self.disease_items[self.current_key]
        disease = item["disease_name"]
        terms = [line.strip() for line in self.term_text.get("1.0", tk.END).splitlines() if line.strip()] if self.term_text else []
        notes = [line.strip() for line in self.notes_text.get("1.0", tk.END).splitlines() if line.strip()] if self.notes_text else []
        query = self.query_text.get("1.0", tk.END).strip() if self.query_text else ""
        comment = self.comment_text.get("1.0", tk.END).strip() if self.comment_text else ""

        payload = {
            "disease_name": disease,
            "canonical_key": item["generated"]["canonical_key"],
            "status": self.strategy_status_var.get().strip() or "pending",
            "generated_on": today_iso(),
            "notes": notes,
            "queries": {
                "openalex_terms": terms,
                "europe_pmc_query": query,
            },
            "review_comment": comment,
        }
        if payload["status"] == "approved":
            payload["approved_on"] = today_iso()
        return payload

    def save_current_item(self) -> None:
        payload = self._build_payload_from_editor()
        if not payload:
            messagebox.showwarning("No Selection", "Select a disease first.")
            return
        self._merge_payload(payload)
        self._save_review_files()
        self.reload_all()
        messagebox.showinfo("Saved", f"Saved strategy for {payload['disease_name']}.")

    def save_all(self) -> None:
        if self.current_key:
            payload = self._build_payload_from_editor()
            if payload:
                self._merge_payload(payload)
        self._save_review_files()
        self.reload_all()
        messagebox.showinfo("Saved", "Saved store and review files.")

    def _merge_payload(self, payload: dict) -> None:
        strategies = self.store.setdefault("strategies", {})
        review_items = self.review_payload.setdefault("review_items", [])
        key = payload["canonical_key"]

        existing_index = next((i for i, item in enumerate(review_items) if item.get("canonical_key") == key), None)
        if existing_index is None:
            review_items.append(payload)
        else:
            review_items[existing_index] = payload

        if payload["status"] == "approved":
            strategies[key] = payload
        elif key in strategies and payload["status"] != "approved":
            del strategies[key]

        self.store["generated_on"] = today_iso()
        self.review_payload["generated_on"] = today_iso()

    def _save_review_files(self) -> None:
        self.review_payload.setdefault("instructions", [
            "Review each pending disease strategy.",
            "Change status from pending to approved or rejected.",
            "You may edit the query terms before approval.",
            "After approval, rerun the retrieval workflow and the approved strategy will move into the local store automatically.",
        ])
        save_json_file(SEARCH_STORE_FILE, self.store)
        save_json_file(SEARCH_REVIEW_FILE, self.review_payload)
        self._write_review_markdown()

    def _write_review_markdown(self) -> None:
        lines = [
            "# Search Strategy Review",
            "",
            "Review the candidate search strategies below.",
            "",
        ]
        for index, item in enumerate(self.review_payload.get("review_items", []), start=1):
            lines.extend(
                [
                    f"## {index}. {item.get('disease_name', '')}",
                    "",
                    f"- Canonical key: `{item.get('canonical_key', '')}`",
                    f"- Status: `{item.get('status', '')}`",
                    "- OpenAlex terms:",
                    f"  - `{', '.join(item.get('queries', {}).get('openalex_terms', []))}`",
                    "- Europe PMC query:",
                    f"  - `{item.get('queries', {}).get('europe_pmc_query', '')}`",
                    "",
                ]
            )
        SEARCH_REVIEW_MARKDOWN.write_text("\n".join(lines), encoding="utf-8")

    def generate_missing(self) -> None:
        pending_items = ensure_strategy_candidates(self.diseases, self.store)
        if not pending_items:
            messagebox.showinfo("No New Diseases", "No missing diseases need a new strategy.")
            return
        existing = {item.get("canonical_key"): item for item in self.review_payload.get("review_items", [])}
        for item in pending_items:
            existing[item["canonical_key"]] = item
        self.review_payload["review_items"] = list(existing.values())
        self.review_payload["generated_on"] = today_iso()
        self._save_review_files()
        self.reload_all()
        messagebox.showinfo("Generated", f"Generated {len(pending_items)} new candidate strategies.")


def main() -> None:
    root = tk.Tk()
    LiteratureReviewApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
