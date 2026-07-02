from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


BASE_DIR = Path(__file__).resolve().parent
DOI_FILE = BASE_DIR / "DOI_list.xlsx"
MARKDOWN_DIR = BASE_DIR / "Markdown"


RESULTS = [
    {
        "Disease": "Age-related Macular Degeneration",
        "Title": "OphthaDT: Generative Digital Twins for Forecasting Visual Acuity Trajectories in Ophthalmology",
        "Publication Date": "2026-06-20",
        "Type": "Preprint",
        "Source Database": "arXiv",
        "Journal/Venue": "arXiv",
        "Authors": "Pietro Belligoli; Nikita Makarov; Sayedali Shetab Boushehri; Fabian Schmich; Raul Rodriguez-Esteban; Michael Menden",
        "DOI": "",
        "Identifier": "arXiv:2606.22101",
        "Open Access": "Yes",
        "License/OA Status": "Open Access",
        "Landing Page": "https://arxiv.org/abs/2606.22101",
        "PDF URL": "https://arxiv.org/pdf/2606.22101",
        "Retrieved-from Record": "Live open-web search verified on 2026-06-30 using disease-name-centered AMD terms",
        "Abstract File": "001_Age-related_Macular_Degeneration_2026-06-20_OphthaDT.md",
        "Abstract": (
            "Precision medicine in ophthalmology requires accurate longitudinal predictions, but the fragmented "
            "nature of multimodal clinical data remains a barrier to forecasting. We introduce OphthaDT, an "
            "LLM-based digital twin for ophthalmology that serializes longitudinal patient histories from 3,220 "
            "patients across four Phase III clinical trials into structured narratives to forecast best corrected "
            "visual acuity (BCVA). In benchmarks spanning up to 100 weeks, OphthaDT demonstrated the lowest "
            "prediction error in neovascular age-related macular degeneration (nAMD), achieving an average mean "
            "absolute error (MAE) reduction of 6.0% compared to all baselines. In diabetic macular edema (DME), "
            "OphthaDT demonstrated competitive performance against all baselines while outperforming Random Forest "
            "and XGBoost by an average MAE reduction of 2.6% and 6.9%, respectively. Results reveal that "
            "OphthaDT's predictive advantage scales with trajectory complexity: whereas linear models remain "
            "effective for the more stable treatment responses of DME, OphthaDT's capacity is better suited for "
            "capturing the high longitudinal variability of nAMD. Finally, OphthaDT handles irregular sampling "
            "without imputation, positioning LLM-based clinical trajectory modeling as a methodology that could "
            "reduce patient burden and accelerate drug development."
        ),
    }
]


NO_RESULT_SUMMARY = [
    {
        "Disease": "Bergmeister Papilla",
        "Status": "No qualifying open-access abstract records found",
        "Date Filter": "After 2026-06-05",
        "Checked Sources": "General web search, arXiv-oriented search, ophthalmology-oriented search",
    }
]


def write_markdown() -> None:
    MARKDOWN_DIR.mkdir(parents=True, exist_ok=True)
    for item in RESULTS:
        path = MARKDOWN_DIR / item["Abstract File"]
        path.write_text(
            "\n".join(
                [
                    f"# {item['Title']}",
                    "",
                    f"- Disease: {item['Disease']}",
                    f"- Publication Date: {item['Publication Date']}",
                    f"- Type: {item['Type']}",
                    f"- Source Database: {item['Source Database']}",
                    f"- Journal/Venue: {item['Journal/Venue']}",
                    f"- Authors: {item['Authors']}",
                    f"- DOI: {item['DOI']}",
                    f"- Identifier: {item['Identifier']}",
                    f"- Open Access: {item['Open Access']}",
                    f"- License/OA Status: {item['License/OA Status']}",
                    f"- Landing Page: {item['Landing Page']}",
                    f"- PDF URL: {item['PDF URL']}",
                    f"- Retrieved-from Record: {item['Retrieved-from Record']}",
                    "",
                    "## Abstract",
                    "",
                    item["Abstract"],
                    "",
                ]
            ),
            encoding="utf-8",
        )


def write_excel() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    headers = [
        "Disease",
        "Title",
        "Publication Date",
        "Type",
        "Source Database",
        "Journal/Venue",
        "Authors",
        "DOI",
        "Identifier",
        "Open Access",
        "License/OA Status",
        "Landing Page",
        "PDF URL",
        "Retrieved-from Record",
        "Abstract File",
    ]
    ws.append(headers)
    for item in RESULTS:
        ws.append([item[h] for h in headers])

    summary = wb.create_sheet("No_Result_Summary")
    summary_headers = ["Disease", "Status", "Date Filter", "Checked Sources"]
    summary.append(summary_headers)
    for item in NO_RESULT_SUMMARY:
        summary.append([item[h] for h in summary_headers])

    meta = wb.create_sheet("Metadata")
    meta.append(["Field", "Value"])
    meta.append(["Generated On", "2026-06-30"])
    meta.append(["Search Cutoff", "After 2026-06-05"])
    meta.append(["Method", "Live open-web search across accessible open-access literature sources using disease-name-centered terms only"])
    meta.append(["Notes", "Results were re-verified from live accessible web sources on 2026-06-30, with the cutoff kept at after 2026-06-05 and without expanding to examination or imaging technique terms."])

    strategy = wb.create_sheet("Search_Strategy")
    strategy.append(["Disease", "Search approach"])
    strategy.append(
        [
            "Age-related Macular Degeneration",
            "Restricted to disease-name-centered terms such as Age-related Macular Degeneration, AMD, age related macular degeneration, neovascular age-related macular degeneration, nAMD, dry AMD, wet AMD, and geographic atrophy when explicitly treated as AMD-related disease expression.",
        ]
    )
    strategy.append(
        [
            "Bergmeister Papilla",
            "Restricted to disease-name-centered terms such as Bergmeister Papilla and closely linked disease expressions including persistent fetal vasculature or persistent hyperplastic primary vitreous only when directly connected to Bergmeister papilla in the literature context.",
        ]
    )

    wb.save(DOI_FILE)


if __name__ == "__main__":
    write_markdown()
    write_excel()
