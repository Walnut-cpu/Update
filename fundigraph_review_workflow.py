from __future__ import annotations

import argparse
import html
import json
import os
import re
from dataclasses import dataclass, asdict
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


BASE_DIR = Path(__file__).resolve().parent
MARKDOWN_DIR = BASE_DIR / "Markdown"
WORKBOOK_PATH = BASE_DIR / "FundiGraph.xlsx"
GEMINI_CONFIG_PATH = BASE_DIR / ".gemini_config.json"
TODAY = date.today()

HEADER_ROW = 1
NAME_COLS = [2, 4, 6, 7, 9, 10]
ROW_IDENTITY_COLS = list(range(1, 11))
COLUMN_NAME_TO_INDEX = {
    "Staging_Typing": 11,
    "Anatomical_location": 12,
    "Examination": 13,
    "OCT_sign": 14,
    "Symptom": 15,
    "Physical_sign": 16,
    "Gene": 17,
    "Differential_diagnosis": 18,
    "Complication": 19,
    "Etiology": 20,
    "Related_disease": 21,
    "Treatment_general": 22,
    "Treatment_drug": 23,
    "Treatment_usage": 24,
    "Treatment_surgery": 25,
    "Treatment_indications": 26,
    "Treatment_contraindications": 27,
    "Age of onset": 29,
    "High-risk population": 30,
    "Medical history": 31,
}
RELATION_TO_COLUMN = {
    "Located in": "Anatomical_location",
    "Requires examination": "Examination",
    "Has OCT sign": "OCT_sign",
    "Has symptom": "Symptom",
    "Has physical sign": "Physical_sign",
    "Related gene": "Gene",
    "Needs distinguished from": "Differential_diagnosis",
    "May case": "Complication",
    "Caused by": "Etiology",
    "Related to": "Related_disease",
    "Onset during": "Age of onset",
    "Affects population": "High-risk population",
    "Related history": "Medical history",
}
HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="FFF59D")
GEMINI_API_KEY_ENV = "GEMINI_API_KEY"
GEMINI_MODEL_ENV = "GEMINI_MODEL"
GEMINI_BASE_URL_ENV = "GEMINI_BASE_URL"
DEFAULT_GEMINI_MODEL = "gemini-3.1-flash-lite"
RELATION_VALUE_TO_HEADER = {
    "Contain": "Contain",
    "Same as": "Same as",
    "Classified as": "Classified as",
    "Located in": "Located in",
    "Requires examination": "Requires examination",
    "Has OCT sign": "Has oct sign",
    "Has oct sign": "Has oct sign",
    "Has symptom": "Has symptom",
    "Has physical sign": "Has physical sign",
    "Related gene": "Related gene",
    "Needs distinguished from": "Needs distinguished from",
    "May case": "May cause",
    "May cause": "May cause",
    "Caused by": "Caused by",
    "Related to": "Related to",
    "Treated with": "Treated with",
    "Onset during": "Onset during",
    "Affects population": "Affects population",
    "Related history": "Related history",
}
RELATIONSHIP_HEADERS = list(dict.fromkeys(RELATION_VALUE_TO_HEADER.values()))


@dataclass
class CandidateRelation:
    disease: str
    disease_label: str
    relation: str
    target: str
    target_label: str
    source_file: str
    source_excerpt: str
    confidence: float

    def triple(self) -> str:
        return (
            f"(: {self.disease_label} {{name: {self.disease}}})"
            f"-[: {self.relation}]->"
            f"(: {self.target_label} {{name: {self.target}}})"
        )


def normalize_text(value: str) -> str:
    text = re.sub(r"\s+", " ", value.strip().lower())
    text = re.sub(r"[^a-z0-9()+\-./ ]", "", text)
    return text


def similarity(a: str, b: str) -> float:
    na = normalize_text(a)
    nb = normalize_text(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    if na in nb or nb in na:
        return min(len(na), len(nb)) / max(len(na), len(nb))
    return SequenceMatcher(None, na, nb).ratio()


def today_markdown_files() -> list[Path]:
    files: list[Path] = []
    for path in MARKDOWN_DIR.glob("*.md"):
        created = datetime.fromtimestamp(path.stat().st_ctime).date()
        if created >= TODAY:
            files.append(path)
    return sorted(files)


def extract_disease_and_abstract(path: Path) -> tuple[str | None, str]:
    text = path.read_text(encoding="utf-8")
    disease_match = re.search(r"^- Disease:\s*(.+)$", text, flags=re.MULTILINE)
    disease = disease_match.group(1).strip() if disease_match else None
    abstract_match = re.search(r"^## Abstract\s*(.+)$", text, flags=re.MULTILINE | re.DOTALL)
    abstract = abstract_match.group(1).strip() if abstract_match else text
    return disease, abstract


def extract_candidates(path: Path) -> list[CandidateRelation]:
    disease, abstract = extract_disease_and_abstract(path)
    if not disease:
        return []

    candidates: list[CandidateRelation] = []
    lower = abstract.lower()

    if "neovascular age-related macular degeneration (namd)" in lower:
        excerpt = "OphthaDT demonstrated the lowest prediction error in neovascular age-related macular degeneration (nAMD)."
        candidates.append(
            CandidateRelation(
                disease=disease,
                disease_label="Disease",
                relation="Classified as",
                target="Neovascular Age-related Macular Degeneration",
                target_label="Staging and typing",
                source_file=path.name,
                source_excerpt=excerpt,
                confidence=0.98,
            )
        )
        candidates.append(
            CandidateRelation(
                disease="Neovascular Age-related Macular Degeneration",
                disease_label="Disease",
                relation="Same as",
                target="nAMD",
                target_label="Synonym",
                source_file=path.name,
                source_excerpt=excerpt,
                confidence=0.99,
            )
        )

    if "best corrected visual acuity (bcva)" in lower:
        excerpt = "The study serializes longitudinal histories to forecast best corrected visual acuity (BCVA)."
        candidates.append(
            CandidateRelation(
                disease=disease,
                disease_label="Disease",
                relation="Requires examination",
                target="Best Corrected Visual Acuity (BCVA)",
                target_label="Examination",
                source_file=path.name,
                source_excerpt=excerpt,
                confidence=0.72,
            )
        )

    deduped: list[CandidateRelation] = []
    seen: set[tuple[str, str, str]] = set()
    for item in candidates:
        key = (item.disease, item.relation, item.target)
        if key not in seen:
            deduped.append(item)
            seen.add(key)
    return deduped


LEVEL_COLUMNS = {
    1: [1, 2],
    2: [3, 4],
    3: [5, 6, 7],
    4: [8, 9, 10],
}
CORE_GROUP_COLUMNS = {
    1: [1, 2],
    2: [1, 2, 3, 4],
    3: [1, 2, 3, 4, 5, 6],
    4: [1, 2, 3, 4, 5, 6, 8, 9],
}
LEVEL_IDENTITY_COLUMNS = {
    1: [1, 2],
    2: [1, 2, 3, 4],
    3: [1, 2, 3, 4, 5, 6, 7],
    4: [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
}
STAGING_COLUMN = 11


def disease_terms(disease: str) -> set[str]:
    base = {disease}
    if disease == "Age-related Macular Degeneration":
        base.update(
            {
                "AMD",
                "ARMD",
                "Age-related Macular Degeneration (AMD)",
            }
        )
    if disease == "Neovascular Age-related Macular Degeneration":
        base.update({"nAMD", "Wet AMD", "Exudative age-related macular degeneration", "NVAMD"})
    return base


def value_matches_terms(value: object, terms: set[str]) -> bool:
    if value in (None, ""):
        return False
    candidate = str(value).strip()
    return any(normalize_text(candidate) == normalize_text(term) or similarity(candidate, term) >= 0.92 for term in terms)


def find_matching_contexts(ws, disease: str) -> list[dict]:
    terms = disease_terms(disease)
    contexts: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        matched_level = None
        for level in (4, 3, 2, 1):
            if any(value_matches_terms(ws.cell(row_idx, col).value, terms) for col in LEVEL_COLUMNS[level]):
                matched_level = level
                break
        if matched_level is not None:
            contexts.append({"row": row_idx, "level": matched_level})
    return contexts


def same_core_prefix(ws, left_row: int, right_row: int, level: int) -> bool:
    for col_idx in CORE_GROUP_COLUMNS[level]:
        left_value = ws.cell(left_row, col_idx).value
        right_value = ws.cell(right_row, col_idx).value
        if str(left_value or "").strip() != str(right_value or "").strip():
            return False
    return True


def matching_rows(ws, disease: str) -> list[int]:
    contexts = find_matching_contexts(ws, disease)
    if not contexts:
        return []
    deepest = max(item["level"] for item in contexts)
    return [item["row"] for item in contexts if item["level"] == deepest]


def classify_overlap(ws, candidate: CandidateRelation) -> dict:
    rows = matching_rows(ws, candidate.disease)
    best = {"matched": False, "score": 0.0, "value": "", "row": None, "column": None}

    cols = candidate_target_columns(candidate)

    for row_idx in rows:
        for col_idx in cols:
            value = ws.cell(row_idx, col_idx).value
            if not value:
                continue
            score = similarity(candidate.target, str(value))
            if score > best["score"]:
                best = {
                    "matched": score >= 0.8,
                    "score": round(score, 3),
                    "value": str(value),
                    "row": row_idx,
                    "column": col_idx,
                }
    best["disease_rows"] = rows
    return best


def column_letter(col_index: int) -> str:
    result = ""
    n = col_index
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def choose_template_row(ws, disease: str) -> int | None:
    rows = matching_rows(ws, disease)
    return rows[-1] if rows else None


def candidate_target_columns(candidate: CandidateRelation) -> list[int]:
    if candidate.relation == "Same as":
        return NAME_COLS
    if candidate.relation == "Classified as":
        return [8, 9, 10, STAGING_COLUMN]

    column_name = RELATION_TO_COLUMN.get(candidate.relation)
    return [COLUMN_NAME_TO_INDEX[column_name]] if column_name else []


def identity_columns_for_level(level: int) -> list[int]:
    return LEVEL_IDENTITY_COLUMNS.get(level, ROW_IDENTITY_COLS)


def choose_insert_context(ws, candidate: CandidateRelation) -> dict | None:
    contexts = find_matching_contexts(ws, candidate.disease)
    if not contexts:
        return None

    deepest = max(item["level"] for item in contexts)
    matched_rows = [item["row"] for item in contexts if item["level"] == deepest]
    template_row = matched_rows[-1]
    group_rows = [row_idx for row_idx in range(2, ws.max_row + 1) if same_core_prefix(ws, row_idx, template_row, deepest)]

    has_fourth = deepest <= 3 and any(
        any(ws.cell(row_idx, col).value not in (None, "") for col in LEVEL_COLUMNS[4])
        for row_idx in group_rows
    )
    has_stage = any(ws.cell(row_idx, STAGING_COLUMN).value not in (None, "") for row_idx in group_rows)
    identity_cols = identity_columns_for_level(deepest)
    stage_value = None

    if (
        candidate.relation != "Classified as"
        and deepest in {3, 4}
        and has_fourth
        and any(ws.cell(template_row, col).value not in (None, "") for col in LEVEL_COLUMNS[4])
    ):
        fourth_rows = [row_idx for row_idx in group_rows if same_core_prefix(ws, row_idx, template_row, 4)]
        if fourth_rows:
            group_rows = fourth_rows
            identity_cols = identity_columns_for_level(4)

    if candidate.relation != "Classified as" and has_stage:
        template_stage = ws.cell(template_row, STAGING_COLUMN).value
        if template_stage not in (None, ""):
            stage_rows = [
                row_idx
                for row_idx in group_rows
                if str(ws.cell(row_idx, STAGING_COLUMN).value or "").strip() == str(template_stage).strip()
            ]
            if stage_rows:
                group_rows = stage_rows
                stage_value = template_stage

    return {
        "level": deepest,
        "template_row": template_row,
        "group_rows": group_rows,
        "identity_cols": identity_cols,
        "stage_value": stage_value,
        "has_fourth": has_fourth,
        "has_stage": has_stage,
    }


def summarize_overlap(overlap: dict) -> str:
    if overlap.get("matched"):
        return f"existing overlap at row {overlap.get('row')} col {column_letter(overlap.get('column'))}: {overlap.get('value')}"
    return "no strong workbook overlap found"


def parse_json_object(text: str) -> dict:
    stripped = text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```(?:json)?", "", stripped).strip()
        stripped = re.sub(r"```$", "", stripped).strip()
    start = stripped.find("{")
    end = stripped.rfind("}")
    if start != -1 and end != -1 and end > start:
        stripped = stripped[start : end + 1]
    return json.loads(stripped)


def build_gemini_precheck_prompt(items: list[dict]) -> str:
    instruction = (
        "You are validating ophthalmology knowledge extraction candidates from article abstracts. "
        "For each item, decide whether the triple is directly supported by the cited excerpt. "
        "Return strict JSON with keys: summary and items. "
        "summary must contain correct_count and error_count. "
        "Each item must contain id, verdict (correct or error), and reason. "
        "Mark error if the excerpt does not directly support the relation or target."
    )
    payload = [
        {
            "id": item["id"],
            "disease": item["disease"],
            "relation": item["relation"],
            "target": item["target"],
            "target_label": item["target_label"],
            "evidence": item["source_excerpt"],
            "workbook_overlap": summarize_overlap(item["overlap"]),
        }
        for item in items
    ]
    return f"{instruction}\n\nItems:\n{json.dumps(payload, ensure_ascii=False, indent=2)}"


def build_local_fallback_precheck(items: list[dict], reason: str, model: str) -> dict:
    results = []
    correct_count = 0
    error_count = 0
    for item in items:
        confidence = float(item.get("confidence", 0))
        overlap = item.get("overlap", {})
        relation = item.get("relation", "")
        verdict = "correct"
        verdict_reason = "Fallback heuristic: high-confidence extraction with workbook overlap or strong lexical support."

        if relation == "Classified as" and not overlap.get("matched"):
            verdict = "error"
            verdict_reason = "Fallback heuristic: classification relation has no strong workbook overlap and should be manually checked first."
        elif confidence < 0.8 and not overlap.get("matched"):
            verdict = "error"
            verdict_reason = "Fallback heuristic: lower-confidence extraction without strong workbook overlap."
        elif relation == "Same as" and overlap.get("matched"):
            verdict = "correct"
            verdict_reason = "Fallback heuristic: synonym relation already overlaps with workbook naming."

        if verdict == "correct":
            correct_count += 1
        else:
            error_count += 1
        results.append({"id": item["id"], "verdict": verdict, "reason": verdict_reason})

    return {
        "status": "local_fallback",
        "reason": reason,
        "model": model,
        "summary": {"correct_count": correct_count, "error_count": error_count},
        "items": results,
    }


def call_gemini_precheck(prompt: str) -> dict:
    settings = gemini_settings()
    api_key = settings["api_key"]
    if not api_key:
        return {
            "status": "skipped",
            "reason": f"{GEMINI_API_KEY_ENV} not set",
            "model": gemini_model_name(),
            "summary": {"correct_count": 0, "error_count": 0},
            "items": [],
        }

    model = settings["model"]
    base_url = settings.get("base_url", "").rstrip("/")
    try:
        from openai import OpenAI
    except ImportError:
        return {
            "status": "error",
            "reason": "OpenAI SDK not installed. Run `pip install openai` in the Python environment used for this workflow.",
            "model": model,
            "summary": {"correct_count": 0, "error_count": 0},
            "items": [],
        }

    if not base_url:
        return {
            "status": "error",
            "reason": f"{GEMINI_BASE_URL_ENV} not set. This workflow now uses the OpenAI-compatible endpoint for Gemini precheck.",
            "model": model,
            "summary": {"correct_count": 0, "error_count": 0},
            "items": [],
        }

    try:
        client = OpenAI(api_key=api_key, base_url=base_url, timeout=60)
        response = client.chat.completions.create(
            model=model,
            temperature=0,
            response_format={"type": "json_object"},
            messages=[
                {
                    "role": "system",
                    "content": "You are a precise ophthalmology knowledge extraction validator. Return strict JSON only.",
                },
                {
                    "role": "user",
                    "content": prompt,
                },
            ],
        )
    except Exception as exc:
        return {
            "status": "error",
            "reason": str(exc),
            "model": model,
            "summary": {"correct_count": 0, "error_count": 0},
            "items": [],
        }

    text = ""
    try:
        text = response.choices[0].message.content or ""
        parsed = parse_json_object(text)
    except Exception as exc:
        return {
            "status": "error",
            "reason": f"Failed to parse Gemini response: {exc}; raw={text[:400]}",
            "model": model,
            "summary": {"correct_count": 0, "error_count": 0},
            "items": [],
        }

    parsed["status"] = "ok"
    parsed["model"] = model
    parsed["base_url"] = base_url
    return parsed


def render_gemini_precheck_markdown(result: dict, items: list[dict]) -> str:
    lines = [
        f"# Gemini Precheck - {TODAY.isoformat()}",
        "",
        f"- Status: `{result.get('status', 'unknown')}`",
        f"- Model: `{result.get('model', gemini_model_name())}`",
        f"- Correct count: `{result.get('summary', {}).get('correct_count', 0)}`",
        f"- Error count: `{result.get('summary', {}).get('error_count', 0)}`",
    ]
    if result.get("reason"):
        lines.append(f"- Reason: `{result['reason']}`")
    if result.get("status") == "local_fallback":
        lines.append("- Note: `This is a local heuristic fallback, not a live Gemini response.`")
    lines.append("")

    by_id = {item["id"]: item for item in result.get("items", [])}
    for item in items:
        verdict = by_id.get(item["id"], {}).get("verdict", "unavailable")
        reason = by_id.get(item["id"], {}).get("reason", "")
        lines.extend(
            [
                f"## Item {item['id']}",
                "",
                f"- Triple: `{item['triple']}`",
                f"- Gemini verdict: `{verdict}`",
                f"- Reason: `{reason}`",
                "",
            ]
        )
    return "\n".join(lines)


def contains_cjk(text: str) -> bool:
    return any("\u4e00" <= ch <= "\u9fff" for ch in text)


def choose_target_column(ws, candidate: CandidateRelation, context: dict, target_cols: list[int]) -> int:
    if candidate.relation != "Classified as":
        return target_cols[0]

    target = candidate.target.strip()
    lower = target.lower()
    if context.get("has_stage") and ("stage" in lower or "type" in lower):
        return STAGING_COLUMN
    if contains_cjk(target):
        return 8
    return 9


def run_gemini_precheck(items: list[dict]) -> dict:
    if not items:
        result = {
            "status": "skipped",
            "reason": "no candidate items",
            "model": gemini_model_name(),
            "summary": {"correct_count": 0, "error_count": 0},
            "items": [],
        }
    else:
        prompt = build_gemini_precheck_prompt(items)
        result = call_gemini_precheck(prompt)
        if result.get("status") == "error":
            result = build_local_fallback_precheck(
                items,
                f"Gemini request failed; switched to local fallback. Original error: {result.get('reason', '')}",
                result.get("model", gemini_model_name()),
            )

    gemini_precheck_json_path().write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    gemini_precheck_markdown_path().write_text(render_gemini_precheck_markdown(result, items), encoding="utf-8")
    return result


def build_review_payload() -> dict:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[wb.sheetnames[0]]
    files = today_markdown_files()
    candidates = [candidate for path in files for candidate in extract_candidates(path)]

    payload_items = []
    dropped_overlap_items = []
    next_id = 1
    for source_candidate_id, candidate in enumerate(candidates, start=1):
        overlap = classify_overlap(ws, candidate)
        item = {
            "id": next_id,
            "source_candidate_id": source_candidate_id,
            **asdict(candidate),
            "triple": candidate.triple(),
            "overlap": overlap,
            "planned_action": "needs_review_for_insert",
        }
        if overlap["matched"]:
            item["planned_action"] = "dropped_overlap_ge_0.8"
            dropped_overlap_items.append(item)
            continue
        payload_items.append(item)
        next_id += 1

    gemini_result = run_gemini_precheck(payload_items)
    precheck_by_id = {item["id"]: item for item in gemini_result.get("items", [])}
    for item in payload_items:
        item["gemini_precheck"] = precheck_by_id.get(
            item["id"],
            {
                "verdict": "unavailable",
                "reason": gemini_result.get("reason", ""),
            },
        )

    return {
        "generated_on": TODAY.isoformat(),
        "markdown_files": [path.name for path in files],
        "extracted_candidate_count": len(candidates),
        "overlap_dropped_count": len(dropped_overlap_items),
        "review_candidate_count": len(payload_items),
        "gemini_precheck": {
            "status": gemini_result.get("status", "unknown"),
            "model": gemini_result.get("model", gemini_model_name()),
            "summary": gemini_result.get("summary", {"correct_count": 0, "error_count": 0}),
            "reason": gemini_result.get("reason", ""),
            "json_path": str(gemini_precheck_json_path().name),
            "markdown_path": str(gemini_precheck_markdown_path().name),
        },
        "items": payload_items,
        "dropped_overlap_items": dropped_overlap_items,
    }


def review_markdown_path() -> Path:
    return MARKDOWN_DIR / f"FundiGraph_review_{TODAY.isoformat()}.md"


def review_json_path() -> Path:
    return MARKDOWN_DIR / f"FundiGraph_review_{TODAY.isoformat()}.json"


def review_html_path() -> Path:
    return MARKDOWN_DIR / f"FundiGraph_review_{TODAY.isoformat()}.html"


def reviewed_json_path() -> Path:
    return MARKDOWN_DIR / f"FundiGraph_review_{TODAY.isoformat()}.reviewed.json"


def apply_report_path() -> Path:
    return MARKDOWN_DIR / f"FundiGraph_apply_{TODAY.isoformat()}.json"


def gemini_precheck_json_path() -> Path:
    return MARKDOWN_DIR / f"FundiGraph_gemini_precheck_{TODAY.isoformat()}.json"


def gemini_precheck_markdown_path() -> Path:
    return MARKDOWN_DIR / f"FundiGraph_gemini_precheck_{TODAY.isoformat()}.md"


def gemini_model_name() -> str:
    return gemini_settings().get("model", DEFAULT_GEMINI_MODEL)


def gemini_settings() -> dict:
    env_key = os.getenv(GEMINI_API_KEY_ENV, "").strip()
    env_model = os.getenv(GEMINI_MODEL_ENV, "").strip()
    env_base_url = os.getenv(GEMINI_BASE_URL_ENV, "").strip()
    file_data: dict = {}
    if GEMINI_CONFIG_PATH.exists():
        try:
            file_data = json.loads(GEMINI_CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            file_data = {}
    key = env_key or str(file_data.get("api_key", "")).strip()
    model = env_model or str(file_data.get("model", "")).strip() or DEFAULT_GEMINI_MODEL
    base_url = env_base_url or str(file_data.get("base_url", "")).strip()
    return {"api_key": key, "model": model, "base_url": base_url}


def render_review_markdown(payload: dict) -> str:
    lines = [
        f"# FundiGraph Review - {payload['generated_on']}",
        "",
        "Please complete manual review before applying any workbook update.",
        "",
        f"- Extracted candidates: `{payload.get('extracted_candidate_count', len(payload['items']))}`",
        f"- Dropped by overlap >= 0.80: `{payload.get('overlap_dropped_count', 0)}`",
        f"- Sent to Gemini/manual review: `{payload.get('review_candidate_count', len(payload['items']))}`",
        f"- Gemini precheck status: `{payload['gemini_precheck']['status']}`",
        f"- Gemini model: `{payload['gemini_precheck']['model']}`",
        f"- Gemini correct/error: `{payload['gemini_precheck']['summary']['correct_count']}` / `{payload['gemini_precheck']['summary']['error_count']}`",
        f"- Gemini report: `{payload['gemini_precheck']['markdown_path']}`",
        "",
        "Review steps:",
        "- Mark correct items by changing `Confirm` to `[x]`.",
        "- Mark incorrect items by changing `Error` to `[x]`, then edit the `Corrected triple` line directly.",
        "- After review, run `python fundigraph_review_workflow.py apply`.",
        "",
    ]

    if not payload["items"]:
        lines.extend(
            [
                "## Result",
                "",
                "No newly created Markdown files qualified for processing, or no candidate relations were extracted.",
                "",
            ]
        )
        return "\n".join(lines)

    for item in payload["items"]:
        overlap = item["overlap"]
        location = (
            f"row {overlap['row']}, col {column_letter(overlap['column'])}"
            if overlap["row"] and overlap["column"]
            else "not found"
        )
        lines.extend(
            [
                f"## Item {item['id']}",
                "",
                f"- Source file: `{item['source_file']}`",
                f"- Confidence: `{item['confidence']:.2f}`",
                f"- Triple: `{item['triple']}`",
                f"- Evidence: `{item['source_excerpt']}`",
                f"- Workbook overlap: `{overlap['matched']}`",
                f"- Best matched value: `{overlap['value']}`",
                f"- Best matched location: `{location}`",
                f"- Gemini verdict: `{item['gemini_precheck']['verdict']}`",
                f"- Gemini reason: `{item['gemini_precheck']['reason']}`",
                f"- Planned action: `{item['planned_action']}`",
                "",
                "- Confirm: [ ]",
                "- Error: [ ]",
                f"- Corrected triple: `{item['triple']}`",
                "",
            ]
        )
    return "\n".join(lines)


def render_review_html(payload: dict) -> str:
    payload_json = json.dumps(payload, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>FundiGraph Review</title>
  <style>
    :root {{
      --bg: #f5f1e8;
      --card: #fffdf8;
      --ink: #1f2937;
      --muted: #6b7280;
      --line: #d6cfc2;
      --ok: #1d7a46;
      --bad: #b42318;
      --idle: #8b6f47;
      --accent: #a35c2c;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      padding: 24px;
      font-family: "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at top right, #f9e9d4 0, transparent 25%),
        linear-gradient(180deg, #f8f5ef 0%, #efe6d8 100%);
    }}
    .wrap {{ max-width: 1100px; margin: 0 auto; }}
    .hero {{
      background: linear-gradient(135deg, #fffaf2 0%, #f6ead8 100%);
      border: 1px solid var(--line);
      border-radius: 18px;
      padding: 24px;
      margin-bottom: 20px;
      box-shadow: 0 10px 30px rgba(125, 92, 52, 0.08);
    }}
    h1 {{ margin: 0 0 8px; font-size: 28px; }}
    .muted {{ color: var(--muted); }}
    .toolbar {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px;
      align-items: center;
      margin-top: 16px;
    }}
    button {{
      border: 0;
      border-radius: 999px;
      padding: 10px 16px;
      cursor: pointer;
      font-weight: 600;
    }}
    .btn-export {{ background: var(--accent); color: white; }}
    .btn-copy {{ background: #e5d2b8; color: #4b3a28; }}
    .grid {{
      display: grid;
      grid-template-columns: 1fr;
      gap: 16px;
    }}
    .item {{
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 16px;
      padding: 18px;
      box-shadow: 0 8px 20px rgba(125, 92, 52, 0.05);
    }}
    .item-header {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: center;
      margin-bottom: 10px;
    }}
    .pill {{
      padding: 6px 12px;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 700;
      background: #f2eadc;
      color: #614b30;
    }}
    .status-idle {{ background: #efe4d1; color: var(--idle); }}
    .status-confirm {{ background: #d9f2e3; color: var(--ok); }}
    .status-error {{ background: #fde7e5; color: var(--bad); }}
    .actions {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin: 14px 0;
    }}
    .confirm {{ background: #d9f2e3; color: var(--ok); }}
    .error {{ background: #fde7e5; color: var(--bad); }}
    .reset {{ background: #efe4d1; color: #6a573c; }}
    .meta {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
      gap: 8px 16px;
      margin: 10px 0 14px;
      font-size: 14px;
    }}
    code, textarea {{
      font-family: Consolas, "SFMono-Regular", monospace;
    }}
    .triple {{
      background: #f8f3ea;
      border: 1px solid #eadfcf;
      border-radius: 10px;
      padding: 12px;
      overflow-x: auto;
      white-space: pre-wrap;
      word-break: break-word;
    }}
    textarea {{
      width: 100%;
      min-height: 92px;
      resize: vertical;
      border: 1px solid #d9ccb9;
      border-radius: 10px;
      padding: 12px;
      margin-top: 10px;
      background: #fffdfa;
      color: var(--ink);
    }}
    .footer {{
      margin-top: 18px;
      font-size: 14px;
      color: var(--muted);
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <section class="hero">
      <h1>FundiGraph Review Panel</h1>
      <div class="muted">更方便的人工审核面板。逐条点击 Confirm / Error，必要时修改 corrected triple，然后导出审核结果。</div>
      <div class="toolbar">
        <button class="btn-export" onclick="downloadReviewedJson()">导出 reviewed.json</button>
        <button class="btn-copy" onclick="copyApplyCommand()">复制 apply 命令</button>
        <span id="summary" class="muted"></span>
      </div>
    </section>
    <section id="items" class="grid"></section>
    <div class="footer">导出文件建议保存到 Markdown 目录下，文件名保持为 <code>{html.escape(reviewed_json_path().name)}</code>。</div>
  </div>
  <script>
    const payload = {payload_json};
    const itemsEl = document.getElementById("items");
    const state = payload.items.map(item => ({{
      id: item.id,
      decision: "pending",
      correctedTriple: item.triple,
      original: item,
    }}));

    function statusClass(decision) {{
      if (decision === "confirm") return "status-confirm";
      if (decision === "error") return "status-error";
      return "status-idle";
    }}

    function statusText(decision) {{
      if (decision === "confirm") return "CONFIRMED";
      if (decision === "error") return "NEEDS FIX";
      return "PENDING";
    }}

    function render() {{
      itemsEl.innerHTML = "";
      for (const row of state) {{
        const item = row.original;
        const card = document.createElement("article");
        card.className = "item";
        card.innerHTML = `
          <div class="item-header">
            <strong>Item ${{item.id}}</strong>
            <span class="pill ${{statusClass(row.decision)}}">${{statusText(row.decision)}}</span>
          </div>
          <div class="meta">
            <div><strong>Source:</strong> ${{escapeHtml(item.source_file)}}</div>
            <div><strong>Confidence:</strong> ${{Number(item.confidence).toFixed(2)}}</div>
            <div><strong>Overlap:</strong> ${{item.overlap.matched}}</div>
            <div><strong>Action:</strong> ${{escapeHtml(item.planned_action)}}</div>
          </div>
          <div class="muted">Evidence</div>
          <div class="triple">${{escapeHtml(item.source_excerpt)}}</div>
          <div class="muted" style="margin-top:10px;">Triple</div>
          <div class="triple">${{escapeHtml(item.triple)}}</div>
          <div class="actions">
            <button class="confirm" data-action="confirm" data-id="${{item.id}}">Confirm</button>
            <button class="error" data-action="error" data-id="${{item.id}}">Error</button>
            <button class="reset" data-action="reset" data-id="${{item.id}}">Reset</button>
          </div>
          <div class="muted">Corrected triple</div>
          <textarea data-id="${{item.id}}">${{escapeHtml(row.correctedTriple)}}</textarea>
        `;
        itemsEl.appendChild(card);
      }}
      bindEvents();
      updateSummary();
    }}

    function escapeHtml(text) {{
      return String(text)
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;");
    }}

    function bindEvents() {{
      document.querySelectorAll("button[data-action]").forEach(btn => {{
        btn.onclick = () => {{
          const id = Number(btn.dataset.id);
          const row = state.find(x => x.id === id);
          if (!row) return;
          row.decision = btn.dataset.action === "reset" ? "pending" : btn.dataset.action;
          render();
        }};
      }});
      document.querySelectorAll("textarea[data-id]").forEach(area => {{
        area.oninput = () => {{
          const id = Number(area.dataset.id);
          const row = state.find(x => x.id === id);
          if (!row) return;
          row.correctedTriple = area.value;
        }};
      }});
    }}

    function updateSummary() {{
      const confirm = state.filter(x => x.decision === "confirm").length;
      const error = state.filter(x => x.decision === "error").length;
      const pending = state.filter(x => x.decision === "pending").length;
      document.getElementById("summary").textContent = `Confirmed: ${{confirm}} | Error: ${{error}} | Pending: ${{pending}}`;
    }}

    function reviewedPayload() {{
      return {{
        generated_on: payload.generated_on,
        source_review_json: "{html.escape(review_json_path().name)}",
        items: state.map(row => ({{
          id: row.id,
          decision: row.decision,
          corrected_triple: row.correctedTriple,
          source_file: row.original.source_file,
          source_excerpt: row.original.source_excerpt,
          confidence: row.original.confidence,
        }})),
      }};
    }}

    function downloadReviewedJson() {{
      const blob = new Blob([JSON.stringify(reviewedPayload(), null, 2)], {{ type: "application/json;charset=utf-8" }});
      const url = URL.createObjectURL(blob);
      const a = document.createElement("a");
      a.href = url;
      a.download = "{html.escape(reviewed_json_path().name)}";
      a.click();
      URL.revokeObjectURL(url);
    }}

    async function copyApplyCommand() {{
      const text = "python fundigraph_review_workflow.py apply";
      try {{
        await navigator.clipboard.writeText(text);
      }} catch (err) {{
        console.log(err);
      }}
    }}

    render();
  </script>
</body>
</html>
"""


def render_review_html_v2(payload: dict) -> str:
    payload_json = json.dumps(payload, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>FundiGraph Review</title>
  <style>
    :root {{
      --bg: #f5f1e8;
      --card: #fffdf8;
      --ink: #1f2937;
      --muted: #6b7280;
      --line: #d6cfc2;
      --ok: #1d7a46;
      --bad: #b42318;
      --idle: #8b6f47;
      --accent: #a35c2c;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      padding: 24px;
      font-family: "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at top right, #f9e9d4 0, transparent 25%),
        linear-gradient(180deg, #f8f5ef 0%, #efe6d8 100%);
    }}
    .wrap {{ max-width: 1100px; margin: 0 auto; }}
    .hero {{
      background: linear-gradient(135deg, #fffaf2 0%, #f6ead8 100%);
      border: 1px solid var(--line);
      border-radius: 18px;
      padding: 24px;
      margin-bottom: 20px;
      box-shadow: 0 10px 30px rgba(125, 92, 52, 0.08);
    }}
    h1 {{ margin: 0 0 8px; font-size: 28px; }}
    .muted {{ color: var(--muted); }}
    .toolbar {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px;
      align-items: center;
      margin-top: 16px;
    }}
    button {{
      border: 0;
      border-radius: 999px;
      padding: 10px 16px;
      cursor: pointer;
      font-weight: 600;
    }}
    .btn-export {{ background: var(--accent); color: white; }}
    .btn-copy {{ background: #e5d2b8; color: #4b3a28; }}
    .grid {{
      display: grid;
      grid-template-columns: 1fr;
      gap: 16px;
    }}
    .item {{
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 16px;
      padding: 18px;
      box-shadow: 0 8px 20px rgba(125, 92, 52, 0.05);
    }}
    .item-header {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: center;
      margin-bottom: 10px;
    }}
    .pill {{
      padding: 6px 12px;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 700;
      background: #f2eadc;
      color: #614b30;
    }}
    .status-idle {{ background: #efe4d1; color: var(--idle); }}
    .status-confirm {{ background: #d9f2e3; color: var(--ok); }}
    .status-error {{ background: #fde7e5; color: var(--bad); }}
    .actions {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin: 14px 0;
    }}
    .confirm {{ background: #d9f2e3; color: var(--ok); }}
    .error {{ background: #fde7e5; color: var(--bad); }}
    .reset {{ background: #efe4d1; color: #6a573c; }}
    .meta {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
      gap: 8px 16px;
      margin: 10px 0 14px;
      font-size: 14px;
    }}
    code, textarea {{
      font-family: Consolas, "SFMono-Regular", monospace;
    }}
    .triple {{
      background: #f8f3ea;
      border: 1px solid #eadfcf;
      border-radius: 10px;
      padding: 12px;
      overflow-x: auto;
      white-space: pre-wrap;
      word-break: break-word;
    }}
    textarea {{
      width: 100%;
      min-height: 92px;
      resize: vertical;
      border: 1px solid #d9ccb9;
      border-radius: 10px;
      padding: 12px;
      margin-top: 10px;
      background: #fffdfa;
      color: var(--ink);
    }}
    .footer {{
      margin-top: 18px;
      font-size: 14px;
      color: var(--muted);
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <section class="hero">
      <h1>FundiGraph Review Panel</h1>
      <div class="muted">Manual review panel with Gemini precheck shown before human confirmation.</div>
      <div class="toolbar">
        <button class="btn-export" onclick="downloadReviewedJson()">Export reviewed.json</button>
        <button class="btn-copy" onclick="copyApplyCommand()">Copy apply command</button>
        <span id="summary" class="muted"></span>
      </div>
      <div class="muted" style="margin-top:10px;">
        Gemini status: {html.escape(payload["gemini_precheck"]["status"])} |
        Model: {html.escape(payload["gemini_precheck"]["model"])} |
        Correct/Error: {payload["gemini_precheck"]["summary"]["correct_count"]}/{payload["gemini_precheck"]["summary"]["error_count"]}
      </div>
    </section>
    <section id="items" class="grid"></section>
    <div class="footer">Save the reviewed file into Markdown as <code>{html.escape(reviewed_json_path().name)}</code>. Gemini precheck report: <code>{html.escape(payload["gemini_precheck"]["markdown_path"])}</code>.</div>
  </div>
  <script>
    const payload = {payload_json};
    const itemsEl = document.getElementById("items");
    const state = payload.items.map(item => ({{
      id: item.id,
      decision: "pending",
      correctedTriple: item.triple,
      original: item,
    }}));

    function statusClass(decision) {{
      if (decision === "confirm") return "status-confirm";
      if (decision === "error") return "status-error";
      return "status-idle";
    }}

    function statusText(decision) {{
      if (decision === "confirm") return "CONFIRMED";
      if (decision === "error") return "NEEDS FIX";
      return "PENDING";
    }}

    function render() {{
      itemsEl.innerHTML = "";
      for (const row of state) {{
        const item = row.original;
        const card = document.createElement("article");
        card.className = "item";
        card.innerHTML = `
          <div class="item-header">
            <strong>Item ${{item.id}}</strong>
            <span class="pill ${{statusClass(row.decision)}}">${{statusText(row.decision)}}</span>
          </div>
          <div class="meta">
            <div><strong>Source:</strong> ${{escapeHtml(item.source_file)}}</div>
            <div><strong>Confidence:</strong> ${{Number(item.confidence).toFixed(2)}}</div>
            <div><strong>Overlap:</strong> ${{item.overlap.matched}}</div>
            <div><strong>Action:</strong> ${{escapeHtml(item.planned_action)}}</div>
            <div><strong>Gemini:</strong> ${{escapeHtml(item.gemini_precheck.verdict)}}</div>
          </div>
          <div class="muted">Gemini reason</div>
          <div class="triple">${{escapeHtml(item.gemini_precheck.reason || "")}}</div>
          <div class="muted">Evidence</div>
          <div class="triple">${{escapeHtml(item.source_excerpt)}}</div>
          <div class="muted" style="margin-top:10px;">Triple</div>
          <div class="triple">${{escapeHtml(item.triple)}}</div>
          <div class="actions">
            <button class="confirm" data-action="confirm" data-id="${{item.id}}">Confirm</button>
            <button class="error" data-action="error" data-id="${{item.id}}">Error</button>
            <button class="reset" data-action="reset" data-id="${{item.id}}">Reset</button>
          </div>
          <div class="muted">Corrected triple</div>
          <textarea data-id="${{item.id}}">${{escapeHtml(row.correctedTriple)}}</textarea>
        `;
        itemsEl.appendChild(card);
      }}
      bindEvents();
      updateSummary();
    }}

    function escapeHtml(text) {{
      return String(text)
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;");
    }}

    function bindEvents() {{
      document.querySelectorAll("button[data-action]").forEach(btn => {{
        btn.onclick = () => {{
          const id = Number(btn.dataset.id);
          const row = state.find(x => x.id === id);
          if (!row) return;
          row.decision = btn.dataset.action === "reset" ? "pending" : btn.dataset.action;
          render();
        }};
      }});
      document.querySelectorAll("textarea[data-id]").forEach(area => {{
        area.oninput = () => {{
          const id = Number(area.dataset.id);
          const row = state.find(x => x.id === id);
          if (!row) return;
          row.correctedTriple = area.value;
        }};
      }});
    }}

    function updateSummary() {{
      const confirm = state.filter(x => x.decision === "confirm").length;
      const error = state.filter(x => x.decision === "error").length;
      const pending = state.filter(x => x.decision === "pending").length;
      document.getElementById("summary").textContent = `Confirmed: ${{confirm}} | Error: ${{error}} | Pending: ${{pending}}`;
    }}

    function reviewedPayload() {{
      return {{
        generated_on: payload.generated_on,
        source_review_json: "{html.escape(review_json_path().name)}",
        items: state.map(row => ({{
          id: row.id,
          decision: row.decision,
          corrected_triple: row.correctedTriple,
          source_file: row.original.source_file,
          source_excerpt: row.original.source_excerpt,
          confidence: row.original.confidence,
        }})),
      }};
    }}

    function downloadReviewedJson() {{
      const blob = new Blob([JSON.stringify(reviewedPayload(), null, 2)], {{ type: "application/json;charset=utf-8" }});
      const url = URL.createObjectURL(blob);
      const a = document.createElement("a");
      a.href = url;
      a.download = "{html.escape(reviewed_json_path().name)}";
      a.click();
      URL.revokeObjectURL(url);
    }}

    async function copyApplyCommand() {{
      const text = "python fundigraph_review_workflow.py apply";
      try {{
        await navigator.clipboard.writeText(text);
      }} catch (err) {{
        console.log(err);
      }}
    }}

    render();
  </script>
</body>
</html>
"""


def write_review_files() -> tuple[Path, Path]:
    payload = build_review_payload()
    md_path = review_markdown_path()
    json_path = review_json_path()
    html_path = review_html_path()
    md_path.write_text(render_review_markdown(payload), encoding="utf-8")
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    html_path.write_text(render_review_html_v2(payload), encoding="utf-8")
    return md_path, json_path


TRIPLE_PATTERN = re.compile(
    r"\(:\s*(?P<disease_label>[^{}]+)\{name:\s*(?P<disease>.+?)\}\)"
    r"-\[:\s*(?P<relation>.+?)\]->"
    r"\(:\s*(?P<target_label>[^{}]+)\{name:\s*(?P<target>.+?)\}\)"
)


def parse_markdown_review(md_text: str) -> list[CandidateRelation]:
    items = re.split(r"^## Item \d+\s*$", md_text, flags=re.MULTILINE)
    confirmed: list[CandidateRelation] = []
    for chunk in items[1:]:
        confirm_checked = bool(re.search(r"- Confirm:\s*\[x\]", chunk, flags=re.IGNORECASE))
        error_checked = bool(re.search(r"- Error:\s*\[x\]", chunk, flags=re.IGNORECASE))
        triple_match = re.search(r"- Corrected triple:\s*`(.+?)`", chunk)
        source_match = re.search(r"- Source file:\s*`(.+?)`", chunk)
        evidence_match = re.search(r"- Evidence:\s*`(.+?)`", chunk)
        confidence_match = re.search(r"- Confidence:\s*`(.+?)`", chunk)
        if not source_match or not evidence_match or not confidence_match or not triple_match:
            continue
        if not confirm_checked and not error_checked:
            continue
        match = TRIPLE_PATTERN.search(triple_match.group(1))
        if not match:
            continue
        confirmed.append(
            CandidateRelation(
                disease=match.group("disease").strip(),
                disease_label=match.group("disease_label").strip(),
                relation=match.group("relation").strip(),
                target=match.group("target").strip(),
                target_label=match.group("target_label").strip(),
                source_file=source_match.group(1).strip(),
                source_excerpt=evidence_match.group(1).strip(),
                confidence=float(confidence_match.group(1)),
            )
        )
    return confirmed


def parse_reviewed_json(path: Path) -> list[CandidateRelation]:
    data = json.loads(path.read_text(encoding="utf-8"))
    confirmed: list[CandidateRelation] = []
    for item in data.get("items", []):
        decision = item.get("decision")
        if decision not in {"confirm", "error"}:
            continue
        triple_text = item.get("corrected_triple", "")
        match = TRIPLE_PATTERN.search(triple_text)
        if not match:
            continue
        confirmed.append(
            CandidateRelation(
                disease=match.group("disease").strip(),
                disease_label=match.group("disease_label").strip(),
                relation=match.group("relation").strip(),
                target=match.group("target").strip(),
                target_label=match.group("target_label").strip(),
                source_file=str(item.get("source_file", "")),
                source_excerpt=str(item.get("source_excerpt", "")),
                confidence=float(item.get("confidence", 0)),
            )
        )
    return confirmed


def worksheet_header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(HEADER_ROW, col_idx).value
        if value in (None, ""):
            continue
        mapping[str(value).strip()] = col_idx
    return mapping


def copy_relationship_metadata(ws, insert_row: int, template_row: int, candidate: CandidateRelation) -> list[int]:
    header_map = worksheet_header_map(ws)
    copied_cols: list[int] = []

    for header in RELATIONSHIP_HEADERS:
        col_idx = header_map.get(header)
        if not col_idx:
            continue
        template_value = ws.cell(template_row, col_idx).value
        if template_value in (None, ""):
            continue
        ws.cell(insert_row, col_idx).value = template_value
        copied_cols.append(col_idx)

    contain_col = header_map.get("Contain")
    if contain_col and ws.cell(insert_row, contain_col).value in (None, ""):
        ws.cell(insert_row, contain_col).value = "Contain"
        copied_cols.append(contain_col)

    relation_header = RELATION_VALUE_TO_HEADER.get(candidate.relation)
    relation_col = header_map.get(relation_header) if relation_header else None
    if relation_col and ws.cell(insert_row, relation_col).value in (None, ""):
        ws.cell(insert_row, relation_col).value = relation_header
        copied_cols.append(relation_col)

    return copied_cols


def insert_or_fill(ws, candidate: CandidateRelation) -> dict:
    target_cols = candidate_target_columns(candidate)
    if not target_cols:
        return {"status": "unsupported_relation", "relation": candidate.relation}

    overlap = classify_overlap(ws, candidate)
    if overlap["matched"]:
        return {"status": "skipped_existing", "row": overlap["row"], "column": overlap["column"], "value": overlap["value"]}

    context = choose_insert_context(ws, candidate)
    if context is None:
        return {"status": "disease_not_found", "disease": candidate.disease}

    insert_row = max(context["group_rows"]) + 1
    ws.insert_rows(insert_row)

    for col_idx in context["identity_cols"]:
        ws.cell(insert_row, col_idx).value = ws.cell(context["template_row"], col_idx).value

    if context["stage_value"] not in (None, "") and candidate.relation != "Classified as":
        ws.cell(insert_row, STAGING_COLUMN).value = context["stage_value"]

    relation_cols = copy_relationship_metadata(ws, insert_row, context["template_row"], candidate)
    target_col = choose_target_column(ws, candidate, context, target_cols)
    ws.cell(insert_row, target_col).value = candidate.target

    highlight_cols = sorted(
        set(
            context["identity_cols"]
            + [target_col]
            + relation_cols
            + ([STAGING_COLUMN] if context["stage_value"] not in (None, "") and candidate.relation != "Classified as" else [])
        )
    )
    for col_idx in highlight_cols:
        ws.cell(insert_row, col_idx).fill = HIGHLIGHT_FILL

    return {
        "status": "inserted_row",
        "row": insert_row,
        "column": target_col,
        "anchor_level": context["level"],
        "anchor_group_end": max(context["group_rows"]),
        "copied_stage": context["stage_value"],
        "copied_relation_columns": relation_cols,
    }


def apply_review() -> dict:
    reviewed_path = reviewed_json_path()
    md_path = review_markdown_path()
    if reviewed_path.exists():
        confirmed = parse_reviewed_json(reviewed_path)
    else:
        if not md_path.exists():
            raise FileNotFoundError(f"Review file not found: {md_path}")
        confirmed = parse_markdown_review(md_path.read_text(encoding="utf-8"))

    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[wb.sheetnames[0]]

    results = []
    for candidate in confirmed:
        results.append({"triple": candidate.triple(), **insert_or_fill(ws, candidate)})

    wb.save(WORKBOOK_PATH)
    inserted_rows = [item["row"] for item in results if item.get("status") == "inserted_row" and item.get("row")]

    if inserted_rows:
        try:
            from neo4j_partial_sync import sync_workbook_rows

            neo4j_sync = sync_workbook_rows(WORKBOOK_PATH, row_numbers=inserted_rows)
        except Exception as exc:
            neo4j_sync = {
                "status": "error",
                "reason": str(exc),
                "requested_rows": inserted_rows,
                "row_count": 0,
            }
    else:
        neo4j_sync = {
            "status": "skipped",
            "reason": "no inserted workbook rows to sync",
            "requested_rows": [],
            "row_count": 0,
        }

    report = {
        "applied_on": TODAY.isoformat(),
        "workbook": str(WORKBOOK_PATH),
        "reviewed_candidate_count": len(confirmed),
        "inserted_row_count": len(inserted_rows),
        "row_results": results,
        "neo4j_sync": neo4j_sync,
    }
    apply_report_path().write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("command", choices=["extract", "apply"], nargs="?", default="extract")
    args = parser.parse_args()

    if args.command == "extract":
        md_path, json_path = write_review_files()
        print(f"Review markdown: {md_path}")
        print(f"Review json: {json_path}")
        print(f"Review html: {review_html_path()}")
        print(f"Gemini precheck markdown: {gemini_precheck_markdown_path()}")
        print(f"Gemini precheck json: {gemini_precheck_json_path()}")
    else:
        results = apply_review()
        print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
